from __future__ import annotations

import hashlib
import hmac
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from urllib.parse import quote_plus

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

from data_dictionary import (
    canonical_pathology_name,
    canonical_province_name,
    canonical_territory_name,
    province_from_territory,
)

st.set_page_config(page_title="Dashboard Call Center", layout="wide", initial_sidebar_state="expanded")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
PICTURES_CALL_CENTER_DIR = Path.home() / "Pictures" / "Call Center"
DEFAULT_APPELS_FILENAME = "APPELS COUSP DU 24 FEVRIER 2026.xlsx"
DEFAULT_ALERTES_FILENAME = "ALERTE COUSP DU 25 FEVRIER 2026.xlsx"
RECORDS_TABLE = "call_center_records"
DASHBOARD_USERS_TABLE = "dashboard_users"
DEFAULT_DB_HOST = "localhost"
DEFAULT_DB_PORT = "5432"
DEFAULT_DB_NAME = "call_center"
DEFAULT_DB_USER = "postgres"
DEFAULT_DB_SCHEMA = "public"

THEME = {
    "teal": "#005f73",
    "teal_dark": "#003f4b",
    "blue": "#1d9bf0",
    "pink": "#ef4b9b",
    "green": "#1f9d55",
    "red": "#dc2626",
    "slate": "#4b5563",
    "light": "#f7fbfc",
}

PROVINCE_COORDS = {
    "Kinshasa": (-4.325, 15.322),
    "Kongo Central": (-5.252, 14.865),
    "Kwango": (-6.200, 17.483),
    "Kwilu": (-5.040, 18.817),
    "Mai-Ndombe": (-2.250, 18.300),
    "Kasai": (-6.000, 21.500),
    "Kasai Central": (-6.150, 23.600),
    "Kasai Oriental": (-6.120, 23.590),
    "Lomami": (-6.140, 24.480),
    "Sankuru": (-2.850, 23.430),
    "Maniema": (-2.950, 26.200),
    "Sud Kivu": (-3.000, 28.100),
    "Nord Kivu": (-0.780, 29.250),
    "Ituri": (1.550, 30.250),
    "Haut-Uele": (3.250, 27.600),
    "Bas-Uele": (3.650, 24.950),
    "Tshopo": (0.520, 25.200),
    "Mongala": (2.300, 21.300),
    "Equateur": (0.100, 19.800),
    "Sud-Ubangi": (3.300, 19.000),
    "Nord-Ubangi": (4.000, 22.000),
    "Tshuapa": (-0.100, 22.700),
    "Tanganyika": (-6.500, 27.450),
    "Haut-Lomami": (-7.800, 24.500),
    "Lualaba": (-10.700, 25.300),
    "Haut-Katanga": (-11.670, 27.480),
}

PROVINCE_ALIASES = {
    "kinshasa": "Kinshasa",
    "kongo central": "Kongo Central",
    "kwango": "Kwango",
    "kwilu": "Kwilu",
    "mai ndombe": "Mai-Ndombe",
    "mai-ndombe": "Mai-Ndombe",
    "kasai": "Kasai",
    "kasai central": "Kasai Central",
    "kasai oriental": "Kasai Oriental",
    "lomami": "Lomami",
    "sankuru": "Sankuru",
    "maniema": "Maniema",
    "sud kivu": "Sud Kivu",
    "nord kivu": "Nord Kivu",
    "ituri": "Ituri",
    "haut uele": "Haut-Uele",
    "bas uele": "Bas-Uele",
    "tshopo": "Tshopo",
    "mongala": "Mongala",
    "equateur": "Equateur",
    "sud ubangi": "Sud-Ubangi",
    "nord ubangi": "Nord-Ubangi",
    "tshuapa": "Tshuapa",
    "tanganyika": "Tanganyika",
    "haut lomami": "Haut-Lomami",
    "lualaba": "Lualaba",
    "haut katanga": "Haut-Katanga",
}

MONTH_TOKEN_MAP = {
    "janvier": 1,
    "janv": 1,
    "january": 1,
    "fevrier": 2,
    "fevr": 2,
    "fev": 2,
    "february": 2,
    "mars": 3,
    "march": 3,
    "avril": 4,
    "avr": 4,
    "april": 4,
    "mai": 5,
    "may": 5,
    "juin": 6,
    "june": 6,
    "juillet": 7,
    "juil": 7,
    "july": 7,
    "aout": 8,
    "august": 8,
    "septembre": 9,
    "sept": 9,
    "september": 9,
    "octobre": 10,
    "oct": 10,
    "october": 10,
    "novembre": 11,
    "nov": 11,
    "november": 11,
    "decembre": 12,
    "dec": 12,
    "december": 12,
}

MONTH_NAMES_FR = {
    1: "Janvier",
    2: "Fevrier",
    3: "Mars",
    4: "Avril",
    5: "Mai",
    6: "Juin",
    7: "Juillet",
    8: "Aout",
    9: "Septembre",
    10: "Octobre",
    11: "Novembre",
    12: "Decembre",
}

CALL_COLUMN_ALIASES = {
    "date": ["date", "periode", "period", "jour", "timestamp"],
    "heure": ["heure", "hour"],
    "province": ["province", "prov"],
    "territoire": ["territoire", "territory", "zone", "district", "ville", "commune"],
    "details": ["details", "detail", "description", "nature", "message", "motif", "sujet", "appel"],
    "incident": ["incident", "pathologie", "maladie", "patho", "type"],
    "categorie": ["categorie", "category", "type demande", "type appel", "classification"],
    "genre": ["genre", "sexe", "gender", "sex"],
    "statut": ["statut", "status", "resolution", "resolu", "conclusion", "etat"],
    "record_count": ["record count", "count", "nombre", "nb", "qty", "quantite", "valeur", "value"],
}

ALERT_COLUMN_ALIASES = {
    "date": ["date", "periode", "period", "mois", "month", "semaine"],
    "heure": ["heure", "hour"],
    "location": ["province", "organisation unit", "organisationunit", "localite", "location", "zone", "territoire"],
    "indicator": ["indicateur", "indicator", "data", "type", "nature", "categorie"],
    "value": ["value", "valeur", "record count", "count", "nombre", "nb", "qty"],
    "details": ["details", "detail", "description", "commentaire", "message"],
}


@dataclass
class DataSourceInfo:
    source: str
    sheet_name: str
    note: str


@dataclass(frozen=True)
class PostgresConfig:
    host: str
    port: str
    database: str
    user: str
    password: str
    schema: str = DEFAULT_DB_SCHEMA
    sslmode: str = "prefer"


def get_secret_or_env(key: str, default: str = "") -> str:
    try:
        if key in st.secrets:
            return str(st.secrets[key])
    except Exception:
        pass
    return str(os.getenv(key, default))


def normalize_dashboard_role(role: str) -> str:
    role_norm = normalize_text(role)
    if role_norm in {"admin", "administrateur", "administrator"}:
        return "administrateur"
    return "utilisateur"


def verify_dashboard_password(raw_password: str, expected: str) -> bool:
    """
    Compare le mot de passe saisi avec la valeur configuree.
    Formats supportes:
    - texte brut: "monmotdepasse"
    - hash sha256: "sha256:<hexdigest>"
    """
    expected = str(expected or "").strip()
    if not expected:
        return False
    if expected.lower().startswith("sha256:"):
        wanted = expected.split(":", 1)[1].strip().lower()
        got = hashlib.sha256((raw_password or "").encode("utf-8")).hexdigest().lower()
        return hmac.compare_digest(got, wanted)
    return hmac.compare_digest(raw_password or "", expected)


def to_password_storage(value: str) -> str:
    """Retourne un mot de passe stockable (sha256:<hash>)."""
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.lower().startswith("sha256:"):
        return raw
    hashed = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return f"sha256:{hashed}"


def ensure_dashboard_users_table(conn_url: str, schema: str) -> str:
    """Cree la table des utilisateurs dashboard si absente et injecte les comptes initiaux."""
    schema_name = sanitize_identifier(schema, DEFAULT_DB_SCHEMA)
    schema_sql = quote_ident(schema_name)
    users_sql = f"{schema_sql}.{quote_ident(DASHBOARD_USERS_TABLE)}"
    engine = get_pg_engine(conn_url)

    with engine.begin() as conn:
        conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {schema_sql}"))
        conn.execute(
            text(
                f"""
                CREATE TABLE IF NOT EXISTS {users_sql} (
                    id BIGSERIAL PRIMARY KEY,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL CHECK (role IN ('administrateur', 'utilisateur')),
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    full_name TEXT,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                )
                """
            )
        )
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{DASHBOARD_USERS_TABLE}_role ON {users_sql} (role)"))
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{DASHBOARD_USERS_TABLE}_active ON {users_sql} (is_active)"))

        existing = conn.execute(text(f"SELECT COUNT(*) FROM {users_sql}")).scalar_one()
        if int(existing or 0) == 0:
            seed_admin_user = get_secret_or_env("DASHBOARD_ADMIN_USERNAME", "admin").strip() or "admin"
            seed_admin_password = to_password_storage(get_secret_or_env("DASHBOARD_ADMIN_PASSWORD", "admin"))
            seed_user_user = get_secret_or_env("DASHBOARD_USER_USERNAME", "user").strip() or "user"
            seed_user_password = to_password_storage(get_secret_or_env("DASHBOARD_USER_PASSWORD", "user"))
            seed_admin_name = get_secret_or_env("DASHBOARD_ADMIN_FULLNAME", "Administrateur")
            seed_user_name = get_secret_or_env("DASHBOARD_USER_FULLNAME", "Utilisateur")

            conn.execute(
                text(
                    f"""
                    INSERT INTO {users_sql} (username, password_hash, role, is_active, full_name)
                    VALUES (:username, :password_hash, :role, TRUE, :full_name)
                    ON CONFLICT (username) DO NOTHING
                    """
                ),
                {
                    "username": seed_admin_user,
                    "password_hash": seed_admin_password,
                    "role": "administrateur",
                    "full_name": seed_admin_name,
                },
            )
            conn.execute(
                text(
                    f"""
                    INSERT INTO {users_sql} (username, password_hash, role, is_active, full_name)
                    VALUES (:username, :password_hash, :role, TRUE, :full_name)
                    ON CONFLICT (username) DO NOTHING
                    """
                ),
                {
                    "username": seed_user_user,
                    "password_hash": seed_user_password,
                    "role": "utilisateur",
                    "full_name": seed_user_name,
                },
            )
    return schema_name


def read_dashboard_users_postgres(cfg: PostgresConfig) -> tuple[dict[str, dict[str, str]], str]:
    """
    Charge les utilisateurs actifs depuis PostgreSQL.
    Retourne (users, source/message).
    """
    conn_url = build_pg_url(cfg)
    schema_name = ensure_dashboard_users_table(conn_url, cfg.schema)
    users_sql = f"{quote_ident(schema_name)}.{quote_ident(DASHBOARD_USERS_TABLE)}"
    query = text(
        f"""
        SELECT username, password_hash, role
        FROM {users_sql}
        WHERE is_active = TRUE
        ORDER BY username
        """
    )
    with get_pg_engine(conn_url).connect() as conn:
        rows = conn.execute(query).mappings().all()

    users: dict[str, dict[str, str]] = {}
    for row in rows:
        username = str(row["username"]).strip()
        password = str(row["password_hash"]).strip()
        role = normalize_dashboard_role(str(row["role"]))
        if username and password:
            users[username] = {"password": password, "role": role}
    return users, f"PostgreSQL {cfg.host}:{cfg.port}/{cfg.database}.{schema_name}.{DASHBOARD_USERS_TABLE}"


def load_dashboard_users_from_secrets() -> tuple[dict[str, dict[str, str]], str]:
    """Fallback secrets/env quand PostgreSQL n'est pas joignable."""
    users: dict[str, dict[str, str]] = {}
    source = "fallback-dev"

    raw_json = get_secret_or_env("DASHBOARD_USERS_JSON", "").strip()
    if raw_json:
        try:
            parsed = json.loads(raw_json)
            if isinstance(parsed, list):
                for row in parsed:
                    if not isinstance(row, dict):
                        continue
                    username = str(row.get("username", "")).strip()
                    password = to_password_storage(str(row.get("password", "")).strip())
                    role = normalize_dashboard_role(str(row.get("role", "utilisateur")))
                    if username and password:
                        users[username] = {"password": password, "role": role}
            source = "DASHBOARD_USERS_JSON"
        except Exception:
            users = {}

    if not users:
        admin_user = get_secret_or_env("DASHBOARD_ADMIN_USERNAME", "").strip()
        admin_password = to_password_storage(get_secret_or_env("DASHBOARD_ADMIN_PASSWORD", "").strip())
        std_user = get_secret_or_env("DASHBOARD_USER_USERNAME", "").strip()
        std_password = to_password_storage(get_secret_or_env("DASHBOARD_USER_PASSWORD", "").strip())
        if admin_user and admin_password:
            users[admin_user] = {"password": admin_password, "role": "administrateur"}
        if std_user and std_password:
            users[std_user] = {"password": std_password, "role": "utilisateur"}
        if users:
            source = "DASHBOARD_*_USERNAME/PASSWORD"

    if not users:
        users = {
            "admin": {"password": to_password_storage("admin"), "role": "administrateur"},
            "user": {"password": to_password_storage("user"), "role": "utilisateur"},
        }
        source = "fallback-dev"
    return users, source


def load_dashboard_users(cfg_override: PostgresConfig | None = None) -> tuple[dict[str, dict[str, str]], str]:
    """
    Charge les utilisateurs.
    Mode par defaut: PostgreSQL.
    Pour activer le fallback auto: DASHBOARD_AUTH_SOURCE=auto
    """
    mode = get_secret_or_env("DASHBOARD_AUTH_SOURCE", "postgres").strip().lower()
    cfg = cfg_override or default_postgres_config()
    if not str(cfg.password or "").strip():
        if mode == "postgres":
            return {}, "postgres-error: mot de passe PostgreSQL manquant (champ 'Mot de passe auth')."
        return load_dashboard_users_from_secrets()
    try:
        ok, db_message = ensure_postgres_database(cfg)
        if not ok:
            raise RuntimeError(db_message)
        users, source = read_dashboard_users_postgres(cfg)
        if users:
            return users, source
        if mode == "postgres":
            return {}, "postgres-error: table dashboard_users vide."
    except Exception as exc:
        if mode == "postgres":
            err = str(exc)
            if "fe_sendauth" in err.lower() and "mot de passe" in err.lower():
                return {}, "postgres-error: mot de passe PostgreSQL invalide ou absent."
            return {}, f"postgres-error: {exc}"

    # Mode auto uniquement.
    return load_dashboard_users_from_secrets()


def list_dashboard_users(cfg: PostgresConfig) -> tuple[pd.DataFrame, str]:
    conn_url = build_pg_url(cfg)
    schema_name = ensure_dashboard_users_table(conn_url, cfg.schema)
    users_sql = f"{quote_ident(schema_name)}.{quote_ident(DASHBOARD_USERS_TABLE)}"
    query = text(
        f"""
        SELECT username, role, is_active, COALESCE(full_name, '') AS full_name, created_at, updated_at
        FROM {users_sql}
        ORDER BY role DESC, username
        """
    )
    with get_pg_engine(conn_url).connect() as conn:
        frame = pd.read_sql_query(query, conn)
    return frame, schema_name


def upsert_dashboard_user(
    cfg: PostgresConfig,
    username: str,
    password: str,
    role: str,
    is_active: bool = True,
    full_name: str = "",
) -> None:
    conn_url = build_pg_url(cfg)
    schema_name = ensure_dashboard_users_table(conn_url, cfg.schema)
    users_sql = f"{quote_ident(schema_name)}.{quote_ident(DASHBOARD_USERS_TABLE)}"
    username = str(username).strip()
    if not username:
        raise ValueError("Nom d'utilisateur vide.")
    pw_hash = to_password_storage(password)
    if not pw_hash:
        raise ValueError("Mot de passe vide.")
    role_value = normalize_dashboard_role(role)

    with get_pg_engine(conn_url).begin() as conn:
        conn.execute(
            text(
                f"""
                INSERT INTO {users_sql} (username, password_hash, role, is_active, full_name, updated_at)
                VALUES (:username, :password_hash, :role, :is_active, :full_name, NOW())
                ON CONFLICT (username) DO UPDATE SET
                    password_hash = EXCLUDED.password_hash,
                    role = EXCLUDED.role,
                    is_active = EXCLUDED.is_active,
                    full_name = EXCLUDED.full_name,
                    updated_at = NOW()
                """
            ),
            {
                "username": username,
                "password_hash": pw_hash,
                "role": role_value,
                "is_active": bool(is_active),
                "full_name": str(full_name or "").strip(),
            },
        )


def set_dashboard_user_active(cfg: PostgresConfig, username: str, is_active: bool) -> None:
    conn_url = build_pg_url(cfg)
    schema_name = ensure_dashboard_users_table(conn_url, cfg.schema)
    users_sql = f"{quote_ident(schema_name)}.{quote_ident(DASHBOARD_USERS_TABLE)}"
    with get_pg_engine(conn_url).begin() as conn:
        conn.execute(
            text(f"UPDATE {users_sql} SET is_active = :is_active, updated_at = NOW() WHERE username = :username"),
            {"is_active": bool(is_active), "username": str(username).strip()},
        )


def render_dashboard_users_admin(cfg: PostgresConfig, current_user: str) -> None:
    """Panneau admin pour gerer les comptes utilisateurs du dashboard."""
    with st.sidebar.expander("Gestion utilisateurs dashboard", expanded=False):
        try:
            ok, db_message = ensure_postgres_database(cfg)
            if not ok:
                raise RuntimeError(db_message)

            users_df, schema_name = list_dashboard_users(cfg)
            st.caption(f"Table: {schema_name}.{DASHBOARD_USERS_TABLE}")
            if users_df.empty:
                st.warning("Aucun utilisateur trouve.")
            else:
                st.dataframe(
                    users_df.rename(
                        columns={
                            "username": "Utilisateur",
                            "role": "Role",
                            "is_active": "Actif",
                            "full_name": "Nom complet",
                            "created_at": "Cree le",
                            "updated_at": "Mis a jour",
                        }
                    ),
                    use_container_width=True,
                    height=200,
                    hide_index=True,
                )

            st.markdown("**Ajouter / mettre a jour un utilisateur**")
            username_new = st.text_input("Nom utilisateur", key="user_admin_username")
            fullname_new = st.text_input("Nom complet", key="user_admin_fullname")
            role_new = st.selectbox("Role", ["utilisateur", "administrateur"], key="user_admin_role")
            password_new = st.text_input("Mot de passe", type="password", key="user_admin_password")
            active_new = st.checkbox("Actif", value=True, key="user_admin_active")
            if st.button("Enregistrer utilisateur", key="user_admin_save_button"):
                if not username_new.strip() or not password_new.strip():
                    st.error("Nom utilisateur et mot de passe obligatoires.")
                else:
                    upsert_dashboard_user(
                        cfg=cfg,
                        username=username_new,
                        password=password_new,
                        role=role_new,
                        is_active=active_new,
                        full_name=fullname_new,
                    )
                    st.success("Utilisateur enregistre.")
                    st.rerun()

            st.markdown("**Activer / desactiver un utilisateur**")
            available_users = users_df["username"].astype(str).tolist() if not users_df.empty else []
            target_user = st.selectbox("Utilisateur cible", options=available_users, key="user_admin_target")
            target_status = st.selectbox("Statut", options=["Activer", "Desactiver"], key="user_admin_target_status")
            if st.button("Appliquer statut", key="user_admin_apply_status"):
                if not target_user:
                    st.error("Selectionnez un utilisateur.")
                elif target_user == current_user and target_status == "Desactiver":
                    st.error("Impossible de desactiver votre propre compte actif.")
                else:
                    set_dashboard_user_active(cfg, target_user, target_status == "Activer")
                    st.success("Statut utilisateur mis a jour.")
                    st.rerun()
        except Exception as exc:
            st.error(f"Gestion utilisateurs indisponible: {exc}")


def render_auth_sidebar() -> tuple[bool, str, str]:
    """Gere la connexion applicative et retourne (ok, username, role)."""
    cfg_default = default_postgres_config()
    with st.sidebar.expander("Configuration auth PostgreSQL", expanded=True):
        st.caption("Renseignez la connexion PostgreSQL utilisee pour l'authentification.")
        auth_host = st.text_input("Host auth", value=cfg_default.host, key="auth_pg_host")
        auth_port = st.text_input("Port auth", value=cfg_default.port, key="auth_pg_port")
        auth_database = st.text_input("Base auth", value=cfg_default.database, key="auth_pg_database")
        auth_user = st.text_input("Utilisateur auth", value=cfg_default.user, key="auth_pg_user")
        auth_password = st.text_input("Mot de passe auth", value=cfg_default.password, type="password", key="auth_pg_password")
        auth_schema = st.text_input("Schema auth", value=cfg_default.schema, key="auth_pg_schema")
        auth_sslmode = st.selectbox(
            "SSL mode auth",
            options=["prefer", "disable", "require", "verify-ca", "verify-full"],
            index=["prefer", "disable", "require", "verify-ca", "verify-full"].index(cfg_default.sslmode)
            if cfg_default.sslmode in {"prefer", "disable", "require", "verify-ca", "verify-full"}
            else 0,
            key="auth_pg_sslmode",
        )

    auth_cfg = PostgresConfig(
        host=auth_host.strip() or DEFAULT_DB_HOST,
        port=auth_port.strip() or DEFAULT_DB_PORT,
        database=auth_database.strip() or DEFAULT_DB_NAME,
        user=auth_user.strip() or DEFAULT_DB_USER,
        password=auth_password,
        schema=sanitize_identifier(auth_schema.strip() or DEFAULT_DB_SCHEMA, DEFAULT_DB_SCHEMA),
        sslmode=auth_sslmode,
    )

    users, source = load_dashboard_users(auth_cfg)
    if source.startswith("postgres-error"):
        st.sidebar.error(f"Authentification PostgreSQL indisponible: {source}")
        st.sidebar.info("Verifiez surtout le mot de passe PostgreSQL (auth), puis la table dashboard_users.")
        return False, "", "utilisateur"
    if source == "fallback-dev":
        st.sidebar.warning("Utilisateurs fallback actifs (admin/admin, user/user). Configurez PostgreSQL en production.")
    elif source.startswith("PostgreSQL"):
        st.sidebar.caption(f"Auth source: {source}")

    auth_ok = bool(st.session_state.get("auth_ok", False))
    auth_user = str(st.session_state.get("auth_user", ""))
    auth_role = str(st.session_state.get("auth_role", "utilisateur"))

    if auth_ok and auth_user:
        st.sidebar.success(f"Connecte: {auth_user} ({auth_role})")
        if st.sidebar.button("Se deconnecter", key="logout_button"):
            st.session_state["auth_ok"] = False
            st.session_state["auth_user"] = ""
            st.session_state["auth_role"] = "utilisateur"
            st.rerun()
        return True, auth_user, auth_role

    with st.sidebar.expander("Connexion utilisateur", expanded=True):
        username = st.text_input("Nom d'utilisateur", key="login_username")
        password = st.text_input("Mot de passe", type="password", key="login_password")
        if st.button("Se connecter", key="login_button"):
            if username not in users:
                st.error("Utilisateur inconnu.")
            else:
                expected = users[username]["password"]
                if verify_dashboard_password(password, expected):
                    role = users[username]["role"]
                    st.session_state["auth_ok"] = True
                    st.session_state["auth_user"] = username
                    st.session_state["auth_role"] = role
                    st.rerun()
                st.error("Mot de passe invalide.")
    st.info("Connectez-vous pour acceder au tableau de bord.")
    return False, "", "utilisateur"


def get_admin_import_secret() -> str:
    return get_secret_or_env("ADMIN_IMPORT_PASSWORD", "").strip()


def render_admin_import_guard() -> bool:
    required_secret = get_admin_import_secret()
    if not required_secret:
        st.caption("Import admin: ouvert (aucun mot de passe admin configure).")
        return True

    if "admin_import_unlocked" not in st.session_state:
        st.session_state["admin_import_unlocked"] = False

    unlocked = bool(st.session_state.get("admin_import_unlocked"))
    if unlocked:
        st.success("Mode administrateur actif: import autorise.")
        if st.button("Verrouiller l'import", key="admin_import_lock_button"):
            st.session_state["admin_import_unlocked"] = False
            st.rerun()
        return True

    code = st.text_input("Code admin import", type="password", key="admin_import_code")
    if st.button("Debloquer l'import", key="admin_import_unlock_button"):
        if hmac.compare_digest(code, required_secret):
            st.session_state["admin_import_unlocked"] = True
            st.rerun()
        st.error("Code admin invalide.")

    st.warning("Import verrouille: lecture seule pour les utilisateurs.")
    return False


def sanitize_identifier(value: str, fallback: str) -> str:
    candidate = re.sub(r"[^A-Za-z0-9_]", "_", str(value or "").strip())
    if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", candidate):
        return candidate.lower()
    return fallback


def quote_ident(identifier: str) -> str:
    return f'"{identifier}"'


def build_pg_url(config: PostgresConfig, database_override: str | None = None) -> str:
    database_name = (database_override or config.database).strip() or DEFAULT_DB_NAME
    user = quote_plus(config.user)
    password = quote_plus(config.password)
    auth = f"{user}:{password}" if config.password else user
    return (
        f"postgresql+psycopg://{auth}@{config.host}:{config.port}/{quote_plus(database_name)}"
        f"?sslmode={config.sslmode}"
    )


@st.cache_resource(show_spinner=False)
def get_pg_engine(conn_url: str) -> Engine:
    return create_engine(conn_url, pool_pre_ping=True)


def is_missing_database_error(error: Exception) -> bool:
    message = str(error).lower()
    english = "database" in message and "does not exist" in message
    french = ("base de donn" in message or "database" in message) and "n'existe pas" in message
    return english or french


def can_auto_create_database(db_name: str) -> bool:
    return re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", db_name or "") is not None


def ensure_postgres_database(config: PostgresConfig) -> tuple[bool, str]:
    db_name = (config.database or "").strip()
    if not db_name:
        return False, "Le nom de la base PostgreSQL est vide."

    target_url = build_pg_url(config)
    target_engine = create_engine(target_url, pool_pre_ping=True)
    try:
        with target_engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True, ""
    except Exception as exc:
        if not is_missing_database_error(exc):
            return False, f"Echec connexion PostgreSQL: {exc}"
    finally:
        target_engine.dispose()

    if not can_auto_create_database(db_name):
        return (
            False,
            f"La base '{db_name}' n'existe pas. Creez-la manuellement dans pgAdmin (nom simple: lettres/chiffres/_).",
        )

    for admin_db in ["postgres", "template1"]:
        if admin_db.lower() == db_name.lower():
            continue
        admin_engine = create_engine(
            build_pg_url(config, database_override=admin_db),
            pool_pre_ping=True,
            isolation_level="AUTOCOMMIT",
        )
        try:
            with admin_engine.connect() as conn:
                exists = conn.execute(
                    text("SELECT 1 FROM pg_database WHERE datname = :db_name"),
                    {"db_name": db_name},
                ).scalar()
                if not exists:
                    conn.execute(text(f'CREATE DATABASE "{db_name}"'))
                    return True, f"Base '{db_name}' creee automatiquement."
                return True, ""
        except Exception:
            continue
        finally:
            admin_engine.dispose()

    return (
        False,
        f"La base '{db_name}' n'existe pas et la creation automatique a echoue. Creez-la dans pgAdmin puis reessayez.",
    )


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return text.strip()


def canonical_province(value: object) -> str:
    raw = "" if pd.isna(value) else str(value).strip()
    if not raw:
        return "Inconnu"

    # First pass: dictionary sourced from national territory/province reference.
    dictionary_match = canonical_province_name(raw)
    if dictionary_match and dictionary_match != raw:
        return dictionary_match

    norm = normalize_text(raw)
    if norm in PROVINCE_ALIASES:
        return PROVINCE_ALIASES[norm]
    for alias, canonical in PROVINCE_ALIASES.items():
        if alias in norm:
            return canonical
    return raw


def normalize_gender(value: object) -> str:
    norm = normalize_text(value)
    if norm in {"h", "homme", "masculin", "male", "m"}:
        return "Homme"
    if norm in {"f", "femme", "feminin", "female"}:
        return "Femme"
    return "ND"


def normalize_status(value: object) -> str:
    norm = normalize_text(value)
    if any(token in norm for token in ["resolu", "resolved", "traite", "cloture", "ferme", "close"]):
        return "Resolu"
    if any(token in norm for token in ["non", "ouvert", "pending", "attente", "cours"]):
        return "Non resolu"
    return "Non resolu"


def normalize_call_details_text(
    table: pd.DataFrame,
    path_col: str = "Pathologie",
    details_col: str = "Details de l'appel",
) -> pd.DataFrame:
    """
    Nettoie la colonne de details pour eviter les doublons visuels avec la pathologie.
    - remplace les valeurs vides / numeriques / identiques a la pathologie
    - produit un texte explicatif lisible pour le tableau final.
    """
    if table is None or table.empty or path_col not in table.columns or details_col not in table.columns:
        return table

    out = table.copy()
    path = out[path_col].fillna("pathologie non precise").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
    details = out[details_col].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()

    numeric_like = details.eq("") | details.str.fullmatch(r"[0-9\+\-\s\(\)\.]{4,}", na=False)
    same_as_pathology = details.map(normalize_text).eq(path.map(normalize_text))
    replace_mask = numeric_like | same_as_pathology

    out[details_col] = details
    out.loc[replace_mask, details_col] = "L'appelant signale des symptomes lies a " + path[replace_mask]
    out.loc[out[details_col].eq(""), details_col] = "Sans detail"
    return out


def is_mostly_numeric_series(series: pd.Series, threshold: float = 0.6) -> bool:
    """
    Detecte les colonnes qui ressemblent a des numeros/identifiants
    (ex: numero d'appel, telephone) pour eviter de les utiliser comme details.
    """
    if series is None or len(series) == 0:
        return False
    text = series.fillna("").astype(str).str.strip()
    text = text[text != ""]
    if text.empty:
        return False
    numeric_like = text.str.fullmatch(r"[0-9\+\-\s\(\)]{6,}", na=False)
    return float(numeric_like.mean()) >= threshold


def choose_details_column(raw: pd.DataFrame, mapping: dict[str, str]) -> str | None:
    """
    Selectionne la meilleure colonne "details":
    - evite les colonnes type numero/id/tel
    - privilegie les colonnes textuelles riches.
    """
    preferred = mapping.get("details")
    if preferred and preferred in raw.columns and not is_mostly_numeric_series(raw[preferred]):
        return preferred

    best_col: str | None = None
    best_score = -1.0
    for col in raw.columns:
        norm = normalize_text(col)
        if not any(tok in norm for tok in ["detail", "description", "nature", "motif", "message", "sujet", "appel"]):
            continue
        if any(tok in norm for tok in ["numero", "num", "id", "tel", "telephone"]):
            continue
        series = raw[col]
        if is_mostly_numeric_series(series):
            continue
        text = series.fillna("").astype(str).str.strip()
        not_empty = text[text != ""]
        if not_empty.empty:
            continue
        avg_len = float(not_empty.str.len().mean())
        unique_ratio = float(not_empty.nunique()) / float(len(not_empty)) if len(not_empty) else 0.0
        score = avg_len + (unique_ratio * 10.0)
        if score > best_score:
            best_score = score
            best_col = col

    return best_col


def clean_label(value: object, fallback: str) -> str:
    text = "" if pd.isna(value) else str(value).strip()
    return text if text else fallback


def make_column_map(df: pd.DataFrame, aliases: dict[str, list[str]]) -> dict[str, str]:
    normalized_cols = {col: normalize_text(col) for col in df.columns}
    selected: dict[str, str] = {}
    used: set[str] = set()

    for target, patterns in aliases.items():
        chosen = None
        for col, norm_col in normalized_cols.items():
            if col in used:
                continue
            if any(pattern in norm_col for pattern in patterns):
                chosen = col
                break
        if chosen:
            selected[target] = chosen
            used.add(chosen)
    return selected


def parse_month_from_label(label: object) -> tuple[pd.Timestamp | None, str]:
    normalized = normalize_text(label)
    if not normalized:
        return None, ""

    year_match = re.search(r"(20\d{2})", normalized)
    if not year_match:
        return None, ""
    year = int(year_match.group(1))

    month_value = None
    month_token = ""
    for token, month_num in MONTH_TOKEN_MAP.items():
        if re.search(rf"\b{re.escape(token)}\b", normalized):
            month_value = month_num
            month_token = token
            break

    if month_value is None:
        return None, ""

    prefix = normalized.split(str(year))[0]
    if month_token:
        prefix = re.sub(rf"\b{re.escape(month_token)}\b", "", prefix).strip()
    prefix = re.sub(r"\s+", " ", prefix).strip(" -_/:")

    try:
        date_value = pd.Timestamp(year=year, month=month_value, day=1)
    except ValueError:
        return None, ""

    return date_value, prefix


def parse_excel_date_series(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    excel_dates = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")

    numeric_mask = numeric.notna()
    invalid_or_suspect = parsed.isna() | (parsed < pd.Timestamp("1990-01-01")) | (parsed > pd.Timestamp("2100-12-31"))
    replace_mask = numeric_mask & invalid_or_suspect
    parsed.loc[replace_mask] = excel_dates.loc[replace_mask]
    return parsed


def parse_time_delta_series(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    delta = pd.Series(pd.NaT, index=series.index, dtype="timedelta64[ns]")

    day_fraction_mask = numeric.notna() & numeric.between(0, 1, inclusive="both")
    if day_fraction_mask.any():
        seconds = (numeric.loc[day_fraction_mask] * 86400).round()
        delta.loc[day_fraction_mask] = pd.to_timedelta(seconds, unit="s")

    hour_mask = numeric.notna() & numeric.between(1, 24, inclusive="left")
    if hour_mask.any():
        seconds = (numeric.loc[hour_mask] * 3600).round()
        delta.loc[hour_mask] = pd.to_timedelta(seconds, unit="s")

    text_delta = pd.to_timedelta(series.astype(str), errors="coerce")
    delta = delta.fillna(text_delta)
    return delta


def parse_datetime_columns(df: pd.DataFrame, date_col: str | None, hour_col: str | None) -> pd.Series:
    if date_col is None or date_col not in df.columns:
        base = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    else:
        base = parse_excel_date_series(df[date_col])

    if hour_col and hour_col in df.columns:
        delta = parse_time_delta_series(df[hour_col]).fillna(pd.Timedelta(0))
        base = base + delta

    return pd.to_datetime(base, errors="coerce")


def _read_best_sheet(source: object) -> tuple[pd.DataFrame, str]:
    xls = pd.ExcelFile(source)
    best_sheet = xls.sheet_names[0]
    best_df = pd.DataFrame()
    best_score = -1.0

    for sheet in xls.sheet_names:
        frame = pd.read_excel(xls, sheet_name=sheet)
        if frame.empty:
            continue

        valid_cols = [c for c in frame.columns if "unnamed" not in normalize_text(c)]
        if not valid_cols:
            valid_cols = list(frame.columns)

        non_empty = 0.0
        if valid_cols:
            non_empty = float(frame[valid_cols].notna().mean().mean())

        score = len(valid_cols) * 1000 + min(len(frame), 5000) + (non_empty * 100)
        if score > best_score:
            best_score = score
            best_sheet = sheet
            best_df = frame

    if best_df.empty:
        best_df = pd.read_excel(xls, sheet_name=best_sheet)

    return best_df, best_sheet


@st.cache_data(show_spinner=False)
def read_excel_best_sheet_from_path(path_str: str) -> tuple[pd.DataFrame, str]:
    return _read_best_sheet(path_str)


@st.cache_data(show_spinner=False)
def read_excel_best_sheet_from_bytes(binary: bytes) -> tuple[pd.DataFrame, str]:
    return _read_best_sheet(BytesIO(binary))


def find_local_excel(kind: str) -> Path | None:
    search_roots = [DATA_DIR, BASE_DIR, BASE_DIR.parent / "MesAnalyses", PICTURES_CALL_CENTER_DIR]
    exact_candidates = {
        "appels": [
            DEFAULT_APPELS_FILENAME,
            "appels.xlsx",
            "appels.xls",
            "appel.xlsx",
            "appel.xls",
        ],
        "alertes": [
            DEFAULT_ALERTES_FILENAME,
            "alertes.xlsx",
            "alertes.xls",
            "alerte.xlsx",
            "alerte.xls",
        ],
    }
    wildcard_candidates = {
        "appels": ["*appels*.xlsx", "*appels*.xls", "*appel*.xlsx", "*appel*.xls"],
        "alertes": ["*alertes*.xlsx", "*alertes*.xls", "*alerte*.xlsx", "*alerte*.xls"],
    }

    for root in search_roots:
        if not root.exists():
            continue
        for filename in exact_candidates[kind]:
            candidate = root / filename
            if candidate.exists() and candidate.is_file():
                return candidate

    ranked: list[Path] = []
    for root in search_roots:
        if not root.exists():
            continue
        for pattern in wildcard_candidates[kind]:
            ranked.extend([p for p in root.glob(pattern) if p.is_file()])

    if not ranked:
        return None

    ranked = sorted(set(ranked), key=lambda p: p.stat().st_mtime, reverse=True)
    return ranked[0]


def build_demo_calls(rows: int = 14000) -> pd.DataFrame:
    rng = np.random.default_rng(7)

    provinces = list(PROVINCE_COORDS.keys())
    incidents = [
        "Signes & Symptomes",
        "Mpox",
        "Generalites",
        "Paludisme",
        "Cholera",
        "Ebola",
        "VIH/SIDA",
        "Sante animale",
        "Tuberculose",
        "Rougeole",
        "Typhoide",
        "Covid-19",
    ]
    categories = [
        "Suggestions /Demandes/ Requetes",
        "Questions/Preoccupations",
        "Alerte",
        "Plaintes/Denonciations/Reclamations",
        "Reconnaissance/Remerciement",
    ]
    detail_templates = [
        "L'appelant voulait avoir le remede contre les maux de tete.",
        "L'appelant voulait avoir le remede contre les maux de ventre.",
        "L'appelant demande des medicaments pour soigner les maux de ventre.",
        "L'appelant dit etre malade.",
        "Signalement d'un cas suspect dans la communaute.",
    ]

    start = np.datetime64("2025-01-01")
    end = np.datetime64("2026-02-15")
    date_values = start + rng.integers(0, (end - start).astype(int), size=rows).astype("timedelta64[D]")

    frame = pd.DataFrame(
        {
            "date": pd.to_datetime(date_values),
            "province": rng.choice(provinces, size=rows),
            "territoire": rng.choice(
                ["Kinshasa", "Lukolela", "Inongo", "Bandalungwa", "Mont Ngafula", "Kabinda", "Goma", "Bukavu"],
                size=rows,
            ),
            "details": rng.choice(detail_templates, size=rows),
            "incident": rng.choice(
                incidents,
                size=rows,
                p=[0.52, 0.19, 0.09, 0.05, 0.025, 0.015, 0.013, 0.01, 0.009, 0.008, 0.006, 0.004],
            ),
            "categorie": rng.choice(categories, size=rows, p=[0.51, 0.38, 0.08, 0.02, 0.01]),
            "genre": rng.choice(["Homme", "Femme", "ND"], size=rows, p=[0.957, 0.042, 0.001]),
            "statut": rng.choice(["Resolu", "Non resolu"], size=rows, p=[0.86, 0.14]),
            "record_count": 1,
        }
    )

    return frame


def standardize_calls(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return pd.DataFrame(
            columns=[
                "date",
                "province",
                "territoire",
                "details",
                "incident",
                "categorie",
                "genre",
                "statut",
                "record_count",
                "source_file",
                "sheet_name",
            ]
        )

    mapping = make_column_map(raw, CALL_COLUMN_ALIASES)
    out = pd.DataFrame(index=raw.index)

    out["date"] = parse_datetime_columns(raw, mapping.get("date"), mapping.get("heure"))

    out["province"] = raw[mapping["province"]] if "province" in mapping else "Inconnu"
    out["territoire"] = raw[mapping["territoire"]] if "territoire" in mapping else out["province"]

    details_col = choose_details_column(raw, mapping)
    if details_col:
        out["details"] = raw[details_col]
    elif "incident" in mapping:
        out["details"] = raw[mapping["incident"]]
    else:
        out["details"] = "Sans detail"

    out["incident"] = raw[mapping["incident"]] if "incident" in mapping else "Non precise"
    out["categorie"] = raw[mapping["categorie"]] if "categorie" in mapping else "Non classe"
    out["genre"] = raw[mapping["genre"]] if "genre" in mapping else "ND"
    out["statut"] = raw[mapping["statut"]] if "statut" in mapping else "Non resolu"

    if "record_count" in mapping:
        out["record_count"] = pd.to_numeric(raw[mapping["record_count"]], errors="coerce").fillna(1)
    else:
        out["record_count"] = 1

    if out["date"].isna().all():
        out["date"] = pd.Timestamp.today().normalize()
    else:
        out["date"] = out["date"].fillna(out["date"].dropna().min())

    out["province"] = out["province"].map(canonical_province)
    out["territoire"] = out["territoire"].map(canonical_territory_name)
    out["territoire"] = out["territoire"].map(lambda x: clean_label(x, "Inconnu"))
    territory_province = out["territoire"].map(province_from_territory)
    out["province"] = np.where(out["province"].eq("Inconnu") & territory_province.notna(), territory_province, out["province"])
    out["details"] = out["details"].map(lambda x: clean_label(x, "Sans detail"))
    out["incident"] = out["incident"].map(canonical_pathology_name)
    out["incident"] = out["incident"].map(lambda x: clean_label(x, "Non precise"))
    out["categorie"] = out["categorie"].map(lambda x: clean_label(x, "Non classe"))
    out["genre"] = out["genre"].map(normalize_gender)
    out["statut"] = out["statut"].map(normalize_status)
    out["record_count"] = pd.to_numeric(out["record_count"], errors="coerce").fillna(1).clip(lower=0)

    # Trace les metadonnees de provenance quand elles existent deja (lecture PostgreSQL).
    if "source_file" in raw.columns:
        out["source_file"] = raw["source_file"].fillna("-").astype(str)
    else:
        out["source_file"] = "-"
    if "sheet_name" in raw.columns:
        out["sheet_name"] = raw["sheet_name"].fillna("-").astype(str)
    else:
        out["sheet_name"] = "-"

    return out


def standardize_alerts(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return pd.DataFrame(columns=["date", "location", "indicator", "value", "details", "source_file", "sheet_name"])

    mapping = make_column_map(raw, ALERT_COLUMN_ALIASES)
    month_columns: list[tuple[str, pd.Timestamp, str]] = []

    for col in raw.columns:
        month_dt, prefix = parse_month_from_label(col)
        if month_dt is not None:
            month_columns.append((col, month_dt, prefix))

    if month_columns:
        value_cols = [col for col, _, _ in month_columns]
        id_vars = [mapping[col] for col in ["location", "indicator", "details"] if col in mapping]

        melted = raw.melt(id_vars=id_vars, value_vars=value_cols, var_name="raw_period", value_name="value")
        melted["value"] = pd.to_numeric(melted["value"], errors="coerce")
        melted = melted.dropna(subset=["value"])

        period_map = {col: dt for col, dt, _ in month_columns}
        prefix_map = {col: pref for col, _, pref in month_columns}

        melted["date"] = melted["raw_period"].map(period_map)
        if "indicator" in mapping:
            melted["indicator"] = melted[mapping["indicator"]]
        else:
            melted["indicator"] = melted["raw_period"].map(prefix_map)

        if "location" in mapping:
            melted["location"] = melted[mapping["location"]]
        else:
            melted["location"] = "Inconnu"

        if "details" in mapping:
            melted["details"] = melted[mapping["details"]]
        else:
            melted["details"] = ""

        out = melted[["date", "location", "indicator", "value", "details"]].copy()
    else:
        out = pd.DataFrame(index=raw.index)
        out["date"] = parse_datetime_columns(raw, mapping.get("date"), mapping.get("heure"))
        out["location"] = raw[mapping["location"]] if "location" in mapping else "Inconnu"
        out["indicator"] = raw[mapping["indicator"]] if "indicator" in mapping else "Alerte"
        if "value" in mapping:
            out["value"] = pd.to_numeric(raw[mapping["value"]], errors="coerce").fillna(1)
        else:
            out["value"] = 1
        out["details"] = raw[mapping["details"]] if "details" in mapping else ""

    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    if out["date"].isna().all():
        out["date"] = pd.Timestamp.today().normalize()
    else:
        out["date"] = out["date"].fillna(out["date"].dropna().min())

    out["location"] = out["location"].map(canonical_territory_name)
    out["location"] = out["location"].map(lambda x: clean_label(x, "Inconnu"))
    out["indicator"] = out["indicator"].map(lambda x: clean_label(x, "Alerte"))
    out["details"] = out["details"].map(lambda x: clean_label(x, ""))
    out["value"] = pd.to_numeric(out["value"], errors="coerce").fillna(0).clip(lower=0)

    # Trace les metadonnees de provenance quand elles existent deja (lecture PostgreSQL).
    if "source_file" in raw.columns:
        out["source_file"] = raw["source_file"].fillna("-").astype(str)
    else:
        out["source_file"] = "-"
    if "sheet_name" in raw.columns:
        out["sheet_name"] = raw["sheet_name"].fillna("-").astype(str)
    else:
        out["sheet_name"] = "-"

    return out


def detect_excel_dataset_kind(raw: pd.DataFrame) -> str:
    """
    Detection souple du type de fichier.
    Si l'entete est identique (cas frequent), on favorise APPELS.
    """
    call_map = make_column_map(raw, CALL_COLUMN_ALIASES)
    alert_map = make_column_map(raw, ALERT_COLUMN_ALIASES)
    call_score = sum(col in call_map for col in ["date", "province", "territoire", "incident", "categorie", "genre", "statut", "record_count"])
    alert_score = sum(col in alert_map for col in ["date", "location", "indicator", "value", "details"])
    return "calls" if call_score >= alert_score else "alerts"


def alerts_to_calls_dataframe(alerts_df: pd.DataFrame) -> pd.DataFrame:
    """Convertit un dataframe ALERTES vers le schema APPELS (table unique)."""
    if alerts_df.empty:
        out = empty_calls_dataframe()
        out["source_kind"] = "alerts"
        return out

    work = alerts_df.copy()
    territory = work["location"].map(canonical_territory_name)
    province_from_loc = work["location"].map(canonical_province)
    province_from_terr = territory.map(province_from_territory)
    province = np.where(province_from_loc != "Inconnu", province_from_loc, province_from_terr.fillna("Inconnu"))

    out = pd.DataFrame(
        {
            "date": pd.to_datetime(work["date"], errors="coerce"),
            "province": province,
            "territoire": territory,
            "details": work["details"].fillna("").astype(str),
            "incident": work["indicator"].map(canonical_pathology_name).fillna("Non precise"),
            "categorie": "Alerte",
            "genre": "ND",
            "statut": "Non resolu",
            "record_count": pd.to_numeric(work["value"], errors="coerce").fillna(0),
            "source_file": work.get("source_file", "-"),
            "sheet_name": work.get("sheet_name", "-"),
            "source_kind": "alerts",
        }
    )
    out["date"] = out["date"].fillna(pd.Timestamp.today().normalize())
    out["province"] = out["province"].map(lambda x: clean_label(x, "Inconnu"))
    out["territoire"] = out["territoire"].map(lambda x: clean_label(x, "Inconnu"))
    out["details"] = out["details"].map(lambda x: clean_label(x, "Sans detail"))
    out["incident"] = out["incident"].map(lambda x: clean_label(x, "Non precise"))
    out["record_count"] = pd.to_numeric(out["record_count"], errors="coerce").fillna(0).clip(lower=0)
    return out


def calls_to_alerts_dataframe(calls_df: pd.DataFrame) -> pd.DataFrame:
    """Derive les alertes depuis la table unique APPELS (categorie alerte ou source alerte)."""
    if calls_df.empty:
        return empty_alerts_dataframe()

    work = calls_df.copy()
    source_kind = work["source_kind"].fillna("").astype(str) if "source_kind" in work.columns else pd.Series("", index=work.index)
    alert_mask = work["categorie"].astype(str).str.contains("alerte", case=False, na=False) | source_kind.eq("alerts")
    if alert_mask.any():
        work = work.loc[alert_mask].copy()
    else:
        return empty_alerts_dataframe()

    location = np.where(
        work["territoire"].astype(str).str.strip().ne("Inconnu"),
        work["territoire"],
        work["province"],
    )
    out = pd.DataFrame(
        {
            "date": pd.to_datetime(work["date"], errors="coerce"),
            "location": pd.Series(location, index=work.index),
            "indicator": work["incident"].map(canonical_pathology_name).fillna("Alerte"),
            "value": pd.to_numeric(work["record_count"], errors="coerce").fillna(0),
            "details": work["details"].fillna("").astype(str),
            "source_file": work.get("source_file", "-"),
            "sheet_name": work.get("sheet_name", "-"),
        }
    )
    out["date"] = out["date"].fillna(pd.Timestamp.today().normalize())
    out["location"] = out["location"].map(canonical_territory_name).map(lambda x: clean_label(x, "Inconnu"))
    out["indicator"] = out["indicator"].map(lambda x: clean_label(x, "Alerte"))
    out["details"] = out["details"].map(lambda x: clean_label(x, ""))
    out["value"] = pd.to_numeric(out["value"], errors="coerce").fillna(0).clip(lower=0)
    return out


def load_unified_data(uploaded_files: list[object] | None) -> tuple[pd.DataFrame, DataSourceInfo, pd.DataFrame, DataSourceInfo]:
    """
    Upload unique multi-fichiers.
    Chaque fichier est normalise puis stocke en schema unique APPELS.
    ALERTES est derive a la volee depuis les lignes alerte/categorie.
    """
    files = [f for f in (uploaded_files or []) if f is not None]
    if not files:
        return (
            empty_calls_dataframe(),
            DataSourceInfo("Aucune source", "-", "Veuillez importer au moins un fichier Excel."),
            empty_alerts_dataframe(),
            DataSourceInfo("Aucune source", "-", "Les alertes seront derivees automatiquement des donnees importees."),
        )

    calls_frames: list[pd.DataFrame] = []
    loaded_files: list[str] = []
    for file_obj in files:
        try:
            frame, sheet = read_excel_best_sheet_from_bytes(file_obj.getvalue())
            kind = detect_excel_dataset_kind(frame)
            if kind == "alerts":
                clean_alerts = standardize_alerts(frame)
                clean_alerts["source_file"] = str(file_obj.name)
                clean_alerts["sheet_name"] = str(sheet)
                clean_calls = alerts_to_calls_dataframe(clean_alerts)
            else:
                clean_calls = standardize_calls(frame)
                clean_calls["source_kind"] = "calls"
            clean_calls["source_file"] = str(file_obj.name)
            clean_calls["sheet_name"] = str(sheet)
            if "source_kind" not in clean_calls.columns:
                clean_calls["source_kind"] = kind
            calls_frames.append(clean_calls)
            loaded_files.append(f"{file_obj.name} ({sheet})")
        except Exception:
            loaded_files.append(f"{file_obj.name} (erreur)")

    if calls_frames:
        calls_df = pd.concat(calls_frames, ignore_index=True)
    else:
        calls_df = empty_calls_dataframe()

    alerts_df = calls_to_alerts_dataframe(calls_df)
    calls_note = f"{len(files)} fichier(s) Excel importe(s) en table unique."
    if loaded_files:
        calls_note += " " + " | ".join(loaded_files[:5])
        if len(loaded_files) > 5:
            calls_note += f" | +{len(loaded_files) - 5} autres"

    calls_info = DataSourceInfo("Upload multiple", "-", calls_note)
    alerts_info = DataSourceInfo(
        "Upload derive depuis APPELS",
        "-",
        f"{format_int(len(alerts_df))} ligne(s) ALERTES derivee(s) automatiquement.",
    )
    return calls_df, calls_info, alerts_df, alerts_info


def empty_calls_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "date",
            "province",
            "territoire",
            "details",
            "incident",
            "categorie",
            "genre",
            "statut",
            "record_count",
            "source_file",
            "sheet_name",
            "source_kind",
        ]
    )


def empty_alerts_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=["date", "location", "indicator", "value", "details", "source_file", "sheet_name"])


def ensure_postgres_tables(conn_url: str, schema: str) -> str:
    schema_name = sanitize_identifier(schema, DEFAULT_DB_SCHEMA)
    schema_sql = quote_ident(schema_name)
    records_sql = f"{schema_sql}.{quote_ident(RECORDS_TABLE)}"
    users_sql = f"{schema_sql}.{quote_ident(DASHBOARD_USERS_TABLE)}"
    engine = get_pg_engine(conn_url)

    with engine.begin() as conn:
        conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {schema_sql}"))
        conn.execute(
            text(
                f"""
                CREATE TABLE IF NOT EXISTS {records_sql} (
                    id BIGSERIAL PRIMARY KEY,
                    date TIMESTAMP,
                    province TEXT,
                    territoire TEXT,
                    details TEXT,
                    incident TEXT,
                    categorie TEXT,
                    genre TEXT,
                    statut TEXT,
                    record_count DOUBLE PRECISION,
                    source_file TEXT,
                    sheet_name TEXT,
                    source_kind TEXT,
                    row_hash TEXT,
                    imported_at TIMESTAMPTZ DEFAULT NOW()
                )
                """
            )
        )
        conn.execute(
            text(
                f"""
                CREATE TABLE IF NOT EXISTS {users_sql} (
                    id BIGSERIAL PRIMARY KEY,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL CHECK (role IN ('administrateur', 'utilisateur')),
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    full_name TEXT,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                )
                """
            )
        )
        conn.execute(text(f"ALTER TABLE {records_sql} ADD COLUMN IF NOT EXISTS source_kind TEXT"))
        conn.execute(text(f"ALTER TABLE {records_sql} ADD COLUMN IF NOT EXISTS row_hash TEXT"))
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{RECORDS_TABLE}_date ON {records_sql} (date)"))
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{RECORDS_TABLE}_province ON {records_sql} (province)"))
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{RECORDS_TABLE}_source_kind ON {records_sql} (source_kind)"))
        conn.execute(text(f"CREATE UNIQUE INDEX IF NOT EXISTS idx_{RECORDS_TABLE}_row_hash ON {records_sql} (row_hash)"))
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{DASHBOARD_USERS_TABLE}_role ON {users_sql} (role)"))
        conn.execute(text(f"CREATE INDEX IF NOT EXISTS idx_{DASHBOARD_USERS_TABLE}_active ON {users_sql} (is_active)"))
        conn.execute(
            text(
                f"""
                CREATE TABLE IF NOT EXISTS {schema_sql}.import_audit (
                    id BIGSERIAL PRIMARY KEY,
                    imported_at TIMESTAMPTZ DEFAULT NOW(),
                    dataset_type TEXT NOT NULL,
                    file_name TEXT NOT NULL,
                    file_name_norm TEXT NOT NULL,
                    file_hash TEXT NOT NULL,
                    sheet_name TEXT,
                    date_min DATE,
                    date_max DATE,
                    total_rows INTEGER DEFAULT 0,
                    rows_inserted INTEGER DEFAULT 0,
                    duplicate_rows INTEGER DEFAULT 0,
                    missing_columns TEXT,
                    missing_rows INTEGER DEFAULT 0,
                    status TEXT NOT NULL,
                    message TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                f"""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_import_audit_dataset_file_hash
                ON {schema_sql}.import_audit (dataset_type, file_hash)
                """
            )
        )
        conn.execute(
            text(
                f"""
                CREATE INDEX IF NOT EXISTS idx_import_audit_dataset_file_period
                ON {schema_sql}.import_audit (dataset_type, file_name_norm, date_min, date_max)
                """
            )
        )

        # Seed initial users if table is empty.
        user_count = conn.execute(text(f"SELECT COUNT(*) FROM {users_sql}")).scalar_one()
        if int(user_count or 0) == 0:
            seed_admin_user = get_secret_or_env("DASHBOARD_ADMIN_USERNAME", "admin").strip() or "admin"
            seed_admin_password = to_password_storage(get_secret_or_env("DASHBOARD_ADMIN_PASSWORD", "admin"))
            seed_user_user = get_secret_or_env("DASHBOARD_USER_USERNAME", "user").strip() or "user"
            seed_user_password = to_password_storage(get_secret_or_env("DASHBOARD_USER_PASSWORD", "user"))
            conn.execute(
                text(
                    f"""
                    INSERT INTO {users_sql} (username, password_hash, role, is_active, full_name)
                    VALUES (:username, :password_hash, :role, TRUE, :full_name)
                    ON CONFLICT (username) DO NOTHING
                    """
                ),
                {
                    "username": seed_admin_user,
                    "password_hash": seed_admin_password,
                    "role": "administrateur",
                    "full_name": "Administrateur",
                },
            )
            conn.execute(
                text(
                    f"""
                    INSERT INTO {users_sql} (username, password_hash, role, is_active, full_name)
                    VALUES (:username, :password_hash, :role, TRUE, :full_name)
                    ON CONFLICT (username) DO NOTHING
                    """
                ),
                {
                    "username": seed_user_user,
                    "password_hash": seed_user_password,
                    "role": "utilisateur",
                    "full_name": "Utilisateur",
                },
            )

    return schema_name


@st.cache_data(show_spinner=False, ttl=30)
def read_calls_from_postgres(conn_url: str, schema: str) -> pd.DataFrame:
    schema_name = sanitize_identifier(schema, DEFAULT_DB_SCHEMA)
    table_sql = f"{quote_ident(schema_name)}.{quote_ident(RECORDS_TABLE)}"
    query = text(
        f"""
        SELECT date, province, territoire, details, incident, categorie, genre, statut, record_count, source_file, sheet_name, source_kind
        FROM {table_sql}
        """
    )
    with get_pg_engine(conn_url).connect() as conn:
        frame = pd.read_sql_query(query, conn)
    if frame.empty:
        return empty_calls_dataframe()
    return standardize_calls(frame)


@st.cache_data(show_spinner=False, ttl=30)
def read_alerts_from_postgres(conn_url: str, schema: str) -> pd.DataFrame:
    calls_df = read_calls_from_postgres(conn_url, schema)
    return calls_to_alerts_dataframe(calls_df)


def load_postgres_data(conn_url: str, schema: str, db_label: str) -> tuple[pd.DataFrame, DataSourceInfo, pd.DataFrame, DataSourceInfo]:
    schema_name = ensure_postgres_tables(conn_url, schema)
    calls_df = read_calls_from_postgres(conn_url, schema_name)
    alerts_df = read_alerts_from_postgres(conn_url, schema_name)
    calls_info = DataSourceInfo(
        source=db_label,
        sheet_name=f"{schema_name}.{RECORDS_TABLE}",
        note=f"{format_int(len(calls_df))} ligne(s) lue(s) depuis la table unique PostgreSQL.",
    )
    alerts_info = DataSourceInfo(
        source=db_label,
        sheet_name=f"{schema_name}.{RECORDS_TABLE} (derive)",
        note=f"{format_int(len(alerts_df))} ligne(s) ALERTES derivee(s) depuis la table unique.",
    )
    return calls_df, calls_info, alerts_df, alerts_info


def compute_file_hash(binary: bytes) -> str:
    """Calcule un hash SHA256 du fichier pour detecter les re-imports."""
    return hashlib.sha256(binary).hexdigest()


def csv_text(values: list[str]) -> str:
    """Transforme une liste en texte CSV lisible pour le rapport."""
    return ", ".join(values) if values else "-"


def detect_missing_rows_calls(df: pd.DataFrame) -> int:
    """Compte les lignes d'appels incompletes selon les champs critiques."""
    if df.empty:
        return 0
    missing_mask = (
        df["province"].astype(str).str.strip().eq("Inconnu")
        | df["territoire"].astype(str).str.strip().eq("Inconnu")
        | df["details"].astype(str).str.strip().eq("Sans detail")
    )
    return int(missing_mask.sum())


def detect_missing_rows_alerts(df: pd.DataFrame) -> int:
    """Compte les lignes d'alertes incompletes selon les champs critiques."""
    if df.empty:
        return 0
    missing_mask = (
        df["location"].astype(str).str.strip().eq("Inconnu")
        | df["indicator"].astype(str).str.strip().eq("Alerte")
    )
    return int(missing_mask.sum())


def normalize_filename(file_name: str) -> str:
    """Normalise le nom du fichier pour les controles de nomenclature."""
    return normalize_text(file_name)


def row_hash_from_columns(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    # Empreinte stable d'une ligne, utilisee pour eviter les doublons a l'import.
    payload = df.copy()
    for col in columns:
        if col == "date":
            payload[col] = pd.to_datetime(payload[col], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
        else:
            payload[col] = payload[col].fillna("").astype(str).str.strip()
    joined = payload[columns].agg("|".join, axis=1)
    return joined.map(lambda x: hashlib.sha256(x.encode("utf-8")).hexdigest())


def was_file_already_imported(
    conn,
    schema_name: str,
    dataset_type: str,
    file_hash: str,
    file_name_norm: str,
    date_min: pd.Timestamp | None,
    date_max: pd.Timestamp | None,
) -> tuple[bool, str]:
    # Controle anti-doublon fichier:
    # 1) meme contenu (hash identique), 2) meme nom normalise + meme periode.
    audit_table = f"{quote_ident(schema_name)}.import_audit"
    by_hash = conn.execute(
        text(
            f"""
            SELECT status
            FROM {audit_table}
            WHERE dataset_type = :dataset_type AND file_hash = :file_hash
            LIMIT 1
            """
        ),
        {"dataset_type": dataset_type, "file_hash": file_hash},
    ).first()
    if by_hash is not None:
        return True, "Doublon detecte: meme contenu de fichier deja importe (hash identique)."

    by_period = conn.execute(
        text(
            f"""
            SELECT status
            FROM {audit_table}
            WHERE dataset_type = :dataset_type
              AND file_name_norm = :file_name_norm
              AND date_min IS NOT DISTINCT FROM :date_min
              AND date_max IS NOT DISTINCT FROM :date_max
            LIMIT 1
            """
        ),
        {
            "dataset_type": dataset_type,
            "file_name_norm": file_name_norm,
            "date_min": date_min.date() if pd.notna(date_min) else None,
            "date_max": date_max.date() if pd.notna(date_max) else None,
        },
    ).first()
    if by_period is not None:
        return True, "Doublon probable: meme nom de fichier et meme periode deja importes."
    return False, ""


def write_import_audit(conn, schema_name: str, report: dict[str, object]) -> None:
    """Ecrit (ou met a jour) le rapport d'import dans la table import_audit."""
    audit_table = f"{quote_ident(schema_name)}.import_audit"
    conn.execute(
        text(
            f"""
            INSERT INTO {audit_table} (
                dataset_type, file_name, file_name_norm, file_hash, sheet_name,
                date_min, date_max, total_rows, rows_inserted, duplicate_rows,
                missing_columns, missing_rows, status, message
            )
            VALUES (
                :dataset_type, :file_name, :file_name_norm, :file_hash, :sheet_name,
                :date_min, :date_max, :total_rows, :rows_inserted, :duplicate_rows,
                :missing_columns, :missing_rows, :status, :message
            )
            ON CONFLICT (dataset_type, file_hash)
            DO UPDATE SET
                imported_at = NOW(),
                date_min = EXCLUDED.date_min,
                date_max = EXCLUDED.date_max,
                total_rows = EXCLUDED.total_rows,
                rows_inserted = EXCLUDED.rows_inserted,
                duplicate_rows = EXCLUDED.duplicate_rows,
                missing_columns = EXCLUDED.missing_columns,
                missing_rows = EXCLUDED.missing_rows,
                status = EXCLUDED.status,
                message = EXCLUDED.message
            """
        ),
        {
            "dataset_type": report["dataset_type"],
            "file_name": report["file_name"],
            "file_name_norm": report["file_name_norm"],
            "file_hash": report["file_hash"],
            "sheet_name": report.get("sheet_name"),
            "date_min": report["date_min"].date() if pd.notna(report.get("date_min")) else None,
            "date_max": report["date_max"].date() if pd.notna(report.get("date_max")) else None,
            "total_rows": int(report.get("total_rows", 0)),
            "rows_inserted": int(report.get("rows_inserted", 0)),
            "duplicate_rows": int(report.get("duplicate_rows", 0)),
            "missing_columns": report.get("missing_columns", "-"),
            "missing_rows": int(report.get("missing_rows", 0)),
            "status": str(report.get("status", "unknown")),
            "message": str(report.get("message", "")),
        },
    )


def prepare_calls_files_for_postgres(uploaded_files: list[object] | None) -> list[dict[str, object]]:
    """Prepare les fichiers APPELS: validation, standardisation, metadonnees et rapport."""
    files = [f for f in (uploaded_files or []) if f is not None]
    prepared: list[dict[str, object]] = []
    expected = ["date", "province", "territoire", "details", "incident", "categorie", "genre", "statut", "record_count"]

    for file_obj in files:
        report: dict[str, object] = {
            "dataset_type": "calls",
            "file_name": str(file_obj.name),
            "file_name_norm": normalize_filename(str(file_obj.name)),
            "file_hash": compute_file_hash(file_obj.getvalue()),
            "sheet_name": "-",
            "status": "error",
            "message": "",
            "total_rows": 0,
            "rows_inserted": 0,
            "duplicate_rows": 0,
            "missing_columns": "-",
            "missing_rows": 0,
            "date_min": pd.NaT,
            "date_max": pd.NaT,
            "data": empty_calls_dataframe(),
        }
        try:
            raw, sheet = read_excel_best_sheet_from_bytes(file_obj.getvalue())
            mapping = make_column_map(raw, CALL_COLUMN_ALIASES)
            report["sheet_name"] = str(sheet)
            missing_columns = [col for col in expected if col not in mapping]
            report["missing_columns"] = csv_text(missing_columns)

            if "date" not in mapping:
                report["status"] = "rejected"
                report["message"] = "Colonne date absente/non reconnue: import refuse."
                prepared.append(report)
                continue

            clean = standardize_calls(raw)
            clean["source_file"] = str(file_obj.name)
            clean["sheet_name"] = str(sheet)
            report["total_rows"] = int(len(clean))
            report["missing_rows"] = detect_missing_rows_calls(clean)
            report["date_min"] = pd.to_datetime(clean["date"], errors="coerce").min()
            report["date_max"] = pd.to_datetime(clean["date"], errors="coerce").max()
            report["status"] = "ready"
            report["message"] = "Fichier prepare."
            report["data"] = clean
        except Exception as exc:
            report["status"] = "error"
            report["message"] = f"Erreur de lecture/standardisation: {exc}"

        prepared.append(report)
    return prepared


def prepare_alerts_files_for_postgres(uploaded_files: list[object] | None) -> list[dict[str, object]]:
    """Prepare les fichiers ALERTES: validation, standardisation, metadonnees et rapport."""
    files = [f for f in (uploaded_files or []) if f is not None]
    prepared: list[dict[str, object]] = []
    expected = ["date", "location", "indicator", "value", "details"]

    for file_obj in files:
        report: dict[str, object] = {
            "dataset_type": "alerts",
            "file_name": str(file_obj.name),
            "file_name_norm": normalize_filename(str(file_obj.name)),
            "file_hash": compute_file_hash(file_obj.getvalue()),
            "sheet_name": "-",
            "status": "error",
            "message": "",
            "total_rows": 0,
            "rows_inserted": 0,
            "duplicate_rows": 0,
            "missing_columns": "-",
            "missing_rows": 0,
            "date_min": pd.NaT,
            "date_max": pd.NaT,
            "data": empty_alerts_dataframe(),
        }
        try:
            raw, sheet = read_excel_best_sheet_from_bytes(file_obj.getvalue())
            mapping = make_column_map(raw, ALERT_COLUMN_ALIASES)
            report["sheet_name"] = str(sheet)
            missing_columns = [col for col in expected if col not in mapping]
            report["missing_columns"] = csv_text(missing_columns)

            if "date" not in mapping:
                report["status"] = "rejected"
                report["message"] = "Colonne date absente/non reconnue: import refuse."
                prepared.append(report)
                continue

            clean = standardize_alerts(raw)
            clean["source_file"] = str(file_obj.name)
            clean["sheet_name"] = str(sheet)
            report["total_rows"] = int(len(clean))
            report["missing_rows"] = detect_missing_rows_alerts(clean)
            report["date_min"] = pd.to_datetime(clean["date"], errors="coerce").min()
            report["date_max"] = pd.to_datetime(clean["date"], errors="coerce").max()
            report["status"] = "ready"
            report["message"] = "Fichier prepare."
            report["data"] = clean
        except Exception as exc:
            report["status"] = "error"
            report["message"] = f"Erreur de lecture/standardisation: {exc}"

        prepared.append(report)
    return prepared


def prepare_unified_files_for_postgres(uploaded_files: list[object] | None) -> list[dict[str, object]]:
    """Prepare des fichiers Excel heterogenes vers le schema unique call_center_records."""
    files = [f for f in (uploaded_files or []) if f is not None]
    prepared: list[dict[str, object]] = []

    for file_obj in files:
        report: dict[str, object] = {
            "dataset_type": "calls",
            "file_name": str(file_obj.name),
            "file_name_norm": normalize_filename(str(file_obj.name)),
            "file_hash": compute_file_hash(file_obj.getvalue()),
            "sheet_name": "-",
            "status": "error",
            "message": "",
            "total_rows": 0,
            "rows_inserted": 0,
            "duplicate_rows": 0,
            "missing_columns": "-",
            "missing_rows": 0,
            "date_min": pd.NaT,
            "date_max": pd.NaT,
            "data": empty_calls_dataframe(),
        }
        try:
            raw, sheet = read_excel_best_sheet_from_bytes(file_obj.getvalue())
            kind = detect_excel_dataset_kind(raw)
            report["dataset_type"] = kind
            report["sheet_name"] = str(sheet)

            if kind == "alerts":
                mapping = make_column_map(raw, ALERT_COLUMN_ALIASES)
                expected = ["date", "location", "indicator", "value", "details"]
                if "date" not in mapping:
                    report["status"] = "rejected"
                    report["message"] = "Colonne date absente/non reconnue: import refuse."
                    prepared.append(report)
                    continue
                clean_alerts = standardize_alerts(raw)
                clean_alerts["source_file"] = str(file_obj.name)
                clean_alerts["sheet_name"] = str(sheet)
                clean = alerts_to_calls_dataframe(clean_alerts)
            else:
                mapping = make_column_map(raw, CALL_COLUMN_ALIASES)
                expected = ["date", "province", "territoire", "details", "incident", "categorie", "genre", "statut", "record_count"]
                if "date" not in mapping:
                    report["status"] = "rejected"
                    report["message"] = "Colonne date absente/non reconnue: import refuse."
                    prepared.append(report)
                    continue
                clean = standardize_calls(raw)
                clean["source_kind"] = "calls"

            report["missing_columns"] = csv_text([col for col in expected if col not in mapping])
            clean["source_file"] = str(file_obj.name)
            clean["sheet_name"] = str(sheet)
            if "source_kind" not in clean.columns:
                clean["source_kind"] = kind

            report["total_rows"] = int(len(clean))
            report["missing_rows"] = detect_missing_rows_calls(clean)
            report["date_min"] = pd.to_datetime(clean["date"], errors="coerce").min()
            report["date_max"] = pd.to_datetime(clean["date"], errors="coerce").max()
            report["status"] = "ready"
            report["message"] = "Fichier prepare (table unique)."
            report["data"] = clean
        except Exception as exc:
            report["status"] = "error"
            report["message"] = f"Erreur de lecture/standardisation: {exc}"
        prepared.append(report)
    return prepared


def write_records_to_postgres(records_df: pd.DataFrame, conn_url: str, schema: str, replace: bool) -> tuple[int, int]:
    """Insere les donnees normalisees dans la table unique call_center_records."""
    if records_df.empty:
        return 0, 0
    schema_name = sanitize_identifier(schema, DEFAULT_DB_SCHEMA)
    table_sql = f"{quote_ident(schema_name)}.{quote_ident(RECORDS_TABLE)}"
    engine = get_pg_engine(conn_url)

    with engine.begin() as conn:
        if replace:
            conn.execute(text(f"TRUNCATE TABLE {table_sql}"))

    work = records_df.copy()
    if "source_kind" not in work.columns:
        work["source_kind"] = "calls"
    work["row_hash"] = row_hash_from_columns(
        work,
        ["date", "province", "territoire", "details", "incident", "categorie", "genre", "statut", "record_count", "source_kind"],
    )
    before_dedup = len(work)
    work = work.drop_duplicates(subset=["row_hash"]).copy()
    internal_duplicates = before_dedup - len(work)

    min_date = pd.to_datetime(work["date"], errors="coerce").min()
    max_date = pd.to_datetime(work["date"], errors="coerce").max()
    existing_hashes: set[str] = set()
    if pd.notna(min_date) and pd.notna(max_date):
        with engine.connect() as conn:
            existing = pd.read_sql_query(
                text(f"SELECT row_hash FROM {table_sql} WHERE date BETWEEN :min_date AND :max_date"),
                conn,
                params={"min_date": min_date, "max_date": max_date},
            )
        existing_hashes = set(existing["row_hash"].dropna().astype(str).tolist())

    work = work.loc[~work["row_hash"].isin(existing_hashes)].copy()
    database_duplicates = (before_dedup - internal_duplicates) - len(work)
    duplicate_rows = internal_duplicates + max(database_duplicates, 0)

    export_cols = [
        "date",
        "province",
        "territoire",
        "details",
        "incident",
        "categorie",
        "genre",
        "statut",
        "record_count",
        "source_file",
        "sheet_name",
        "source_kind",
        "row_hash",
    ]
    for col in export_cols:
        if col not in work.columns:
            work[col] = None
    if not work.empty:
        work[export_cols].to_sql(
            RECORDS_TABLE,
            engine,
            schema=schema_name,
            if_exists="append",
            index=False,
            method="multi",
            chunksize=2000,
        )
    return int(len(work)), int(duplicate_rows)


def import_uploaded_excels_to_postgres(
    conn_url: str,
    schema: str,
    excel_files: list[object] | None,
    write_mode: str,
) -> tuple[str, pd.DataFrame]:
    """Import Excel multi-fichiers vers table unique PostgreSQL."""
    schema_name = ensure_postgres_tables(conn_url, schema)
    prepared = prepare_unified_files_for_postgres(excel_files)

    if not prepared:
        return "Aucun fichier selectionne.", pd.DataFrame()

    replace_mode = write_mode == "replace"
    engine = get_pg_engine(conn_url)
    records_truncated = False
    report_rows: list[dict[str, object]] = []

    for file_report in prepared:
        report = {k: v for k, v in file_report.items() if k != "data"}
        dataset_type = str(report["dataset_type"])
        file_df = file_report["data"]

        if report["status"] != "ready":
            report["rows_inserted"] = 0
            report["duplicate_rows"] = 0
            with engine.begin() as conn:
                write_import_audit(conn, schema_name, report)
            report_rows.append(report)
            continue

        with engine.begin() as conn:
            if not replace_mode:
                duplicate_file, duplicate_message = was_file_already_imported(
                    conn=conn,
                    schema_name=schema_name,
                    dataset_type=dataset_type,
                    file_hash=str(report["file_hash"]),
                    file_name_norm=str(report["file_name_norm"]),
                    date_min=report["date_min"],
                    date_max=report["date_max"],
                )
                if duplicate_file:
                    report["status"] = "duplicate_file"
                    report["message"] = duplicate_message
                    report["rows_inserted"] = 0
                    report["duplicate_rows"] = int(report.get("total_rows", 0))
                    write_import_audit(conn, schema_name, report)
                    report_rows.append(report)
                    continue

        should_replace = replace_mode and not records_truncated
        inserted_rows, duplicate_rows = write_records_to_postgres(
            records_df=file_df,
            conn_url=conn_url,
            schema=schema_name,
            replace=should_replace,
        )
        records_truncated = records_truncated or should_replace

        report["rows_inserted"] = inserted_rows
        report["duplicate_rows"] = duplicate_rows
        if inserted_rows > 0:
            report["status"] = "imported"
            report["message"] = "Import termine (table unique)."
        else:
            report["status"] = "duplicate_rows"
            report["message"] = "Aucune nouvelle ligne inseree (doublons detectes)."

        with engine.begin() as conn:
            write_import_audit(conn, schema_name, report)
        report_rows.append(report)

    read_calls_from_postgres.clear()
    read_alerts_from_postgres.clear()

    report_df = pd.DataFrame(report_rows)
    total_inserted = int(report_df.get("rows_inserted", pd.Series(dtype=int)).fillna(0).sum()) if not report_df.empty else 0
    total_duplicates = int(report_df.get("duplicate_rows", pd.Series(dtype=int)).fillna(0).sum()) if not report_df.empty else 0
    summary = (
        f"Import termine ({schema_name}.{RECORDS_TABLE}) - Lignes inserees: {format_int(total_inserted)} | "
        f"Doublons ignores: {format_int(total_duplicates)} | Fichiers traites: {format_int(len(report_rows))}."
    )
    return summary, report_df


def format_int(value: float) -> str:
    return f"{int(round(value)):,}".replace(",", " ")


def add_bar_value_labels(fig, orientation: str = "v", is_percent: bool = False, expand_axis: bool = True) -> None:
    if orientation == "h":
        template = "%{x:.1f}%" if is_percent else "%{x:,.0f}"
        # Pour les barres horizontales, on force le texte a l'exterieur pour eviter
        # la disparition des valeurs sur certaines configurations d'affichage.
        text_position = "outside"
    else:
        template = "%{y:.1f}%" if is_percent else "%{y:,.0f}"
        text_position = "outside"
    fig.update_traces(
        texttemplate=template,
        textposition=text_position,
        cliponaxis=False,
        textfont={"color": "#f8fafc", "size": 14},
    )
    if expand_axis and orientation == "h":
        max_x = 0.0
        for trace in fig.data:
            try:
                x_vals = pd.to_numeric(pd.Series(trace.x), errors="coerce")
                trace_max = float(x_vals.max()) if not x_vals.empty else 0.0
                max_x = max(max_x, trace_max)
            except Exception:
                continue
        if max_x > 0:
            fig.update_xaxes(range=[0, max_x * 1.14], automargin=True)
    # "show" garde les valeurs visibles meme quand l'espace est reduit.
    fig.update_layout(uniformtext_minsize=1, uniformtext_mode="show")


def add_line_value_labels(fig, is_percent: bool = False) -> None:
    template = "%{y:.1f}%" if is_percent else "%{y:,.0f}"
    fig.update_traces(mode="lines+markers+text", texttemplate=template, textposition="top center")
    fig.update_layout(uniformtext_minsize=7, uniformtext_mode="hide")


def add_line_end_labels(
    fig,
    trend_df: pd.DataFrame,
    group_col: str,
    value_col: str = "value",
    color: str = "#f8fafc",
) -> None:
    """Ajoute une etiquette de valeur uniquement au dernier point de chaque serie."""
    if trend_df.empty or group_col not in trend_df.columns:
        return
    latest = (
        trend_df.sort_values("date")
        .groupby(group_col, as_index=False)
        .tail(1)
        .copy()
    )
    if latest.empty:
        return
    latest[value_col] = pd.to_numeric(latest[value_col], errors="coerce").fillna(0)
    fig.add_scatter(
        x=latest["date"],
        y=latest[value_col],
        mode="text",
        text=latest[value_col].round().astype(int).astype(str),
        textposition="top center",
        textfont={"color": color, "size": 13},
        showlegend=False,
        hoverinfo="skip",
    )


def inject_styles() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;600;700&family=Rajdhani:wght@600;700&display=swap');

        .stApp {
            background: radial-gradient(circle at top left, #0b121a 0%, #0f1724 55%, #111827 100%);
            font-family: 'Barlow', sans-serif;
            color: #e7eef6;
        }
        .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6, .stApp p, .stApp label {
            color: #e7eef6;
        }
        .main .block-container {
            padding-top: 1rem;
            padding-bottom: 1rem;
        }
        .cc-header {
            display: flex;
            align-items: center;
            gap: 12px;
            background: linear-gradient(90deg, #005f73 0%, #00758d 60%, #005f73 100%);
            color: #ffffff;
            border: 2px solid #0d2f38;
            border-radius: 8px;
            padding: 0.75rem 1rem;
            margin-bottom: 0.75rem;
            animation: ccFadeIn 450ms ease-out;
        }
        .cc-header .logo {
            background: #ffffff;
            color: #003f4b;
            border-radius: 6px;
            font-weight: 700;
            padding: 6px 10px;
            font-size: 0.86rem;
        }
        .cc-header .title {
            font-family: 'Rajdhani', sans-serif;
            letter-spacing: 0.02em;
            font-size: 1.55rem;
            font-weight: 700;
        }
        .kpi-wrap {
            background: transparent;
            padding: 0.25rem 0;
            animation: ccRise 420ms ease-out;
        }
        .kpi-pill {
            height: 34px;
            background: #e9eef1;
            border: 1px solid #afbec7;
            border-radius: 14px;
            margin-bottom: 0.6rem;
        }
        .kpi-card {
            border-radius: 14px;
            border: 3px solid #0d3b52;
            color: #ffffff;
            padding: 0.45rem 0.55rem 0.55rem 0.55rem;
            text-align: center;
            margin-bottom: 0.65rem;
            box-shadow: 0 2px 0 rgba(9, 40, 56, 0.4);
        }
        .kpi-label {
            font-size: 1.02rem;
            font-weight: 700;
            line-height: 1.1;
        }
        .kpi-value {
            font-family: 'Rajdhani', sans-serif;
            font-size: 2.2rem;
            font-weight: 700;
            line-height: 1;
            margin-top: 0.25rem;
        }
        .kpi-horizontal-wrap {
            margin: 0.2rem 0 0.7rem 0;
            padding: 0.55rem 0.55rem 0.25rem 0.55rem;
            border: 1px solid #1f3f52;
            background: #0f2230;
            border-radius: 12px;
        }
        .kpi-horizontal-grid {
            display: grid;
            grid-template-columns: repeat(7, minmax(130px, 1fr));
            gap: 0.45rem;
            overflow-x: auto;
            padding-bottom: 0.2rem;
        }
        .kpi-horizontal-card {
            border-radius: 14px;
            border: 2px solid #103a4f;
            color: #ffffff;
            text-align: center;
            padding: 0.25rem 0.45rem 0.35rem 0.45rem;
            min-height: 72px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            box-shadow: 0 2px 0 rgba(9, 40, 56, 0.35);
        }
        .kpi-horizontal-label {
            font-size: 0.95rem;
            font-weight: 700;
            line-height: 1.05;
        }
        .kpi-horizontal-value {
            font-family: 'Rajdhani', sans-serif;
            font-size: 2rem;
            font-weight: 700;
            line-height: 1;
            margin-top: 0.18rem;
        }
        .source-note {
            border-left: 4px solid #005f73;
            background: #13202c;
            padding: 0.45rem 0.65rem;
            margin: 0.3rem 0;
            border-radius: 6px;
            font-size: 0.88rem;
            color: #d8e6f5;
        }
        .upload-banner {
            border: 1px dashed #0d677b;
            background: #0f2230;
            border-radius: 8px;
            padding: 0.75rem 0.9rem;
            margin: 0.35rem 0 0.65rem 0;
            color: #d8e6f5;
            font-size: 0.92rem;
        }
        .filter-title {
            background: #0f2d3a;
            border: 1px solid #1f4a5c;
            border-radius: 8px;
            color: #e8f5ff;
            padding: 0.45rem 0.65rem;
            margin-bottom: 0.55rem;
            font-weight: 700;
            font-size: 0.96rem;
        }
        @keyframes ccFadeIn {
            from { opacity: 0; transform: translateY(-8px); }
            to { opacity: 1; transform: translateY(0); }
        }
        @keyframes ccRise {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_header() -> None:
    st.markdown(
        """
        <div class="cc-header">
            <div class="logo">CALL CENTER</div>
            <div class="title">DASHBOARD CALL CENTER</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_source_notes(calls_info: DataSourceInfo, alerts_info: DataSourceInfo) -> None:
    st.markdown(
        f"<div class='source-note'><b>Appels</b>: {calls_info.note}<br><b>Source</b>: {calls_info.source} | <b>Feuille</b>: {calls_info.sheet_name}</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<div class='source-note'><b>Alertes</b>: {alerts_info.note}<br><b>Source</b>: {alerts_info.source} | <b>Feuille</b>: {alerts_info.sheet_name}</div>",
        unsafe_allow_html=True,
    )


def render_kpi_card(label: str, value: float, color: str) -> str:
    return (
        f"<div class='kpi-card' style='background:{color};'>"
        f"<div class='kpi-label'>{label}</div>"
        f"<div class='kpi-value'>{format_int(value)}</div>"
        f"</div>"
    )


def compute_kpis(filtered: pd.DataFrame) -> dict[str, float]:
    total_calls = filtered["record_count"].sum()
    provinces_count = filtered["province"].nunique()
    resolved_calls = filtered.loc[filtered["statut"] == "Resolu", "record_count"].sum()
    unresolved_calls = max(total_calls - resolved_calls, 0)
    male_calls = filtered.loc[filtered["genre"] == "Homme", "record_count"].sum()
    female_calls = filtered.loc[filtered["genre"] == "Femme", "record_count"].sum()
    nd_calls = filtered.loc[filtered["genre"] == "ND", "record_count"].sum()
    return {
        "provinces_count": provinces_count,
        "total_calls": total_calls,
        "resolved_calls": resolved_calls,
        "unresolved_calls": unresolved_calls,
        "male_calls": male_calls,
        "female_calls": female_calls,
        "nd_calls": nd_calls,
    }


def render_kpi_panel(kpis: dict[str, float]) -> None:
    st.markdown("<div class='kpi-wrap'>", unsafe_allow_html=True)
    st.markdown("<div class='kpi-pill'></div>", unsafe_allow_html=True)
    st.markdown(render_kpi_card("Province", kpis["provinces_count"], "#1d95ab"), unsafe_allow_html=True)
    st.markdown(render_kpi_card("Tot appels", kpis["total_calls"], "#2d80c1"), unsafe_allow_html=True)
    st.markdown(render_kpi_card("Resolus", kpis["resolved_calls"], "#23a153"), unsafe_allow_html=True)
    st.markdown(render_kpi_card("Non resolus", kpis["unresolved_calls"], "#e61f25"), unsafe_allow_html=True)
    st.markdown(render_kpi_card("Hommes", kpis["male_calls"], "#2f95da"), unsafe_allow_html=True)
    st.markdown(render_kpi_card("Femmes", kpis["female_calls"], "#e04a98"), unsafe_allow_html=True)
    st.markdown(render_kpi_card("ND", kpis["nd_calls"], "#909db0"), unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


def render_kpi_horizontal(kpis: dict[str, float]) -> None:
    cards = [
        ("Province", kpis["provinces_count"], "#1d95ab"),
        ("Tot appels", kpis["total_calls"], "#2d80c1"),
        ("Resolu", kpis["resolved_calls"], "#23a153"),
        ("Non resolu", kpis["unresolved_calls"], "#e61f25"),
        ("Hommes", kpis["male_calls"], "#2f95da"),
        ("Femmes", kpis["female_calls"], "#e04a98"),
        ("ND", kpis["nd_calls"], "#909db0"),
    ]
    html_parts = ["<div class='kpi-horizontal-wrap'><div class='kpi-horizontal-grid'>"]
    for label, value, color in cards:
        html_parts.append(
            f"<div class='kpi-horizontal-card' style='background:{color};'>"
            f"<div class='kpi-horizontal-label'>{label}</div>"
            f"<div class='kpi-horizontal-value'>{format_int(value)}</div>"
            "</div>"
        )
    html_parts.append("</div></div>")
    st.markdown("".join(html_parts), unsafe_allow_html=True)


def group_by_day(df: pd.DataFrame, value_col: str, category_col: str | None = None) -> pd.DataFrame:
    frame = df.copy()
    frame["__day"] = pd.to_datetime(frame["date"], errors="coerce").dt.floor("D")
    frame = frame.dropna(subset=["__day"])
    if frame.empty:
        columns = ["date", value_col] if category_col is None else ["date", category_col, value_col]
        return pd.DataFrame(columns=columns)

    group_cols = ["__day"] + ([category_col] if category_col else [])
    out = (
        frame.groupby(group_cols, as_index=False)[value_col]
        .sum()
        .rename(columns={"__day": "date"})
        .sort_values("date")
    )
    return out


def format_month_option(month_number: int) -> str:
    """Libelle mois pour les filtres reactifs (ex: 02 - Fevrier)."""
    month_num = int(month_number)
    return f"{month_num:02d} - {MONTH_NAMES_FR.get(month_num, f'Mois {month_num}')}"


def build_temporal_filter_frame(df: pd.DataFrame) -> pd.DataFrame:
    """
    Construit les dimensions temporelles utilisees par la zone de filtres:
    annee, mois, semaine ISO et jour.
    """
    if df.empty or "date" not in df.columns:
        return pd.DataFrame(columns=["jour", "annee", "mois", "semaine"])

    date_values = pd.to_datetime(df["date"], errors="coerce")
    out = pd.DataFrame({"jour": date_values.dt.floor("D")})
    out = out.dropna(subset=["jour"])
    if out.empty:
        return pd.DataFrame(columns=["jour", "annee", "mois", "semaine"])

    out["annee"] = out["jour"].dt.year.astype(int)
    out["mois"] = out["jour"].dt.month.astype(int)
    out["semaine"] = out["jour"].dt.isocalendar().week.astype(int)
    return out


def apply_temporal_mask(
    date_values: pd.Series,
    years: list[int] | None,
    months: list[int] | None,
    iso_weeks: list[int] | None,
    days: list[object] | None,
) -> pd.Series:
    """Applique les filtres annee/mois/semaine/jour sur une serie datetime."""
    mask = pd.Series(True, index=date_values.index, dtype=bool)

    if years:
        year_values = [int(v) for v in years]
        mask &= date_values.dt.year.isin(year_values)
    if months:
        month_values = [int(v) for v in months]
        mask &= date_values.dt.month.isin(month_values)
    if iso_weeks:
        week_values = [int(v) for v in iso_weeks]
        mask &= date_values.dt.isocalendar().week.astype("Int64").isin(week_values)
    if days:
        day_values = pd.to_datetime(pd.Series(days), errors="coerce").dt.date.dropna().unique().tolist()
        if day_values:
            mask &= date_values.dt.date.isin(day_values)

    return mask


def apply_calls_filters(
    calls_df: pd.DataFrame,
    date_start: pd.Timestamp,
    date_end: pd.Timestamp,
    provinces: list[str],
    genres: list[str],
    incidents: list[str],
    categories: list[str],
    years: list[int] | None = None,
    months: list[int] | None = None,
    iso_weeks: list[int] | None = None,
    days: list[object] | None = None,
) -> pd.DataFrame:
    """
    Filtre les appels selon:
    - plage de dates
    - dimensions metier
    - granularite temporelle reactive (annee/mois/semaine/jour)
    """
    date_values = pd.to_datetime(calls_df["date"], errors="coerce")
    mask = (date_values >= date_start) & (date_values <= date_end)

    if provinces:
        mask &= calls_df["province"].isin(provinces)
    if genres:
        mask &= calls_df["genre"].isin(genres)
    if incidents:
        mask &= calls_df["incident"].isin(incidents)
    if categories:
        mask &= calls_df["categorie"].isin(categories)

    mask &= apply_temporal_mask(date_values, years, months, iso_weeks, days)
    return calls_df.loc[mask].copy()


def apply_alerts_filters(
    alerts_df: pd.DataFrame,
    date_start: pd.Timestamp,
    date_end: pd.Timestamp,
    provinces: list[str],
    years: list[int] | None = None,
    months: list[int] | None = None,
    iso_weeks: list[int] | None = None,
    days: list[object] | None = None,
) -> pd.DataFrame:
    """Filtre les alertes par periode, granularite temporelle et province (via territoire)."""
    if alerts_df.empty:
        return alerts_df.copy()

    date_values = pd.to_datetime(alerts_df["date"], errors="coerce")
    mask = (date_values >= date_start) & (date_values <= date_end)
    mask &= apply_temporal_mask(date_values, years, months, iso_weeks, days)
    out = alerts_df.loc[mask].copy()

    if provinces:
        wanted = {canonical_province(p) for p in provinces}
        location_province = out["location"].map(canonical_province)
        location_territory = out["location"].map(canonical_territory_name)
        territory_province = location_territory.map(province_from_territory)
        keep_mask = location_province.isin(wanted) | territory_province.isin(wanted)
        out = out.loc[keep_mask].copy()
    return out


def resolve_single_selected_province(filtered_calls: pd.DataFrame, selected_provinces: list[str]) -> str | None:
    """Retourne la province unique selectionnee (sinon None)."""
    selected_clean = [canonical_province(p) for p in selected_provinces if str(p).strip()]
    if len(selected_clean) == 1:
        return selected_clean[0]
    unique_in_data = sorted(filtered_calls["province"].dropna().astype(str).unique().tolist())
    if len(unique_in_data) == 1:
        return unique_in_data[0]
    return None


def compute_completeness_table(df: pd.DataFrame) -> tuple[pd.DataFrame, float]:
    fields = [
        ("date", "Date"),
        ("province", "Province"),
        ("territoire", "Territoire"),
        ("genre", "Genre"),
        ("categorie", "Categorie d'appel"),
        ("incident", "Incident/Pathologie"),
        ("details", "Details de l'appel"),
        ("statut", "Statut"),
    ]
    missing_tokens = {"", "nan", "none", "inconnu", "nd", "non precise", "non classe", "sans detail"}
    sheet_col = "sheet_name" if "sheet_name" in df.columns else None
    rows = []

    for col, label in fields:
        if col not in df.columns:
            continue
        series = df[col]
        valid = series.notna()
        string_series = series.astype(str).str.strip().str.lower()
        valid &= ~string_series.isin(missing_tokens)
        missing_mask = ~valid
        total = int(len(series))
        filled = int(valid.sum())
        rate = (filled / total * 100.0) if total else 0.0

        # Feedback qualite: indique les feuilles qui contiennent des lignes vides pour ce champ.
        missing_sheet_text = "-"
        if sheet_col and missing_mask.any():
            missing_sheets = (
                df.loc[missing_mask, sheet_col]
                .fillna("-")
                .astype(str)
                .str.strip()
                .replace("", "-")
                .unique()
                .tolist()
            )
            missing_sheets = sorted([sheet for sheet in missing_sheets if sheet != "-"])
            if missing_sheets:
                if len(missing_sheets) > 6:
                    missing_sheet_text = ", ".join(missing_sheets[:6]) + f" (+{len(missing_sheets) - 6})"
                else:
                    missing_sheet_text = ", ".join(missing_sheets)

        rows.append(
            {
                "Champ": label,
                "Renseigne": filled,
                "Total": total,
                "Completude (%)": round(rate, 1),
                "Feuilles avec vides": missing_sheet_text,
            }
        )

    if not rows:
        return pd.DataFrame(columns=["Champ", "Renseigne", "Total", "Completude (%)", "Feuilles avec vides"]), 0.0

    table = pd.DataFrame(rows)
    overall = float(table["Renseigne"].sum()) / float(table["Total"].sum()) * 100.0 if table["Total"].sum() else 0.0
    return table, round(overall, 1)


def completeness_level(rate: float) -> str:
    """Classe de completude pour appliquer une couleur de seuil metier."""
    if rate < 15:
        return "< 15%"
    if rate < 50:
        return "15 - 49%"
    if rate <= 90:
        return "50 - 90%"
    return "> 90%"


def build_missing_feedback_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Produit un rapport detaille des valeurs manquantes:
    No, Nom du fichier, Champ vide, Nb lignes vides (par feuille).
    """
    if df.empty:
        return pd.DataFrame(columns=["No", "Nom du fichier", "Champ vide", "Nb lignes vides"])

    fields = [
        ("date", "Date"),
        ("province", "Province"),
        ("territoire", "Territoire"),
        ("genre", "Genre"),
        ("categorie", "Categorie d'appel"),
        ("incident", "Incident/Pathologie"),
        ("details", "Details de l'appel"),
        ("statut", "Statut"),
    ]
    missing_tokens = {"", "nan", "none", "inconnu", "nd", "non precise", "non classe", "sans detail"}
    has_file = "source_file" in df.columns
    has_sheet = "sheet_name" in df.columns
    rows: list[dict[str, object]] = []

    for col, label in fields:
        if col not in df.columns:
            continue
        series = df[col]
        valid = series.notna()
        valid &= ~series.astype(str).str.strip().str.lower().isin(missing_tokens)
        missing_mask = ~valid
        if not missing_mask.any():
            continue

        miss_df = df.loc[missing_mask].copy()
        if has_file:
            miss_df["__file"] = miss_df["source_file"].fillna("-").astype(str).str.strip().replace("", "-")
        else:
            miss_df["__file"] = "-"
        if has_sheet:
            miss_df["__sheet"] = miss_df["sheet_name"].fillna("-").astype(str).str.strip().replace("", "-")
        else:
            miss_df["__sheet"] = "-"
        miss_df["Nom du fichier"] = miss_df["__file"] + " | " + miss_df["__sheet"]

        grouped = (
            miss_df.groupby("Nom du fichier", as_index=False)
            .size()
            .rename(columns={"size": "Nb lignes vides"})
        )
        grouped["Champ vide"] = label
        rows.extend(grouped[["Nom du fichier", "Champ vide", "Nb lignes vides"]].to_dict("records"))

    if not rows:
        return pd.DataFrame(columns=["No", "Nom du fichier", "Champ vide", "Nb lignes vides"])

    feedback = pd.DataFrame(rows).sort_values(["Nom du fichier", "Champ vide"]).reset_index(drop=True)
    feedback.insert(0, "No", feedback.index + 1)
    feedback["Nb lignes vides"] = pd.to_numeric(feedback["Nb lignes vides"], errors="coerce").fillna(0).astype(int)
    return feedback


def render_interactive_analytics(
    filtered_calls: pd.DataFrame,
    previous_calls: pd.DataFrame,
    date_start: pd.Timestamp,
    date_end: pd.Timestamp,
) -> None:
    st.markdown("<div class='filter-title'>Analyses avancees interactives</div>", unsafe_allow_html=True)
    options = st.multiselect(
        "Choisissez les analyses a afficher",
        options=[
            "Qualite des donnees (completude)",
            "Performance de resolution",
            "Comparaison periode precedente",
            "Profil des appels",
        ],
        default=[
            "Qualite des donnees (completude)",
            "Performance de resolution",
            "Comparaison periode precedente",
        ],
        key="analysis_options",
    )

    if filtered_calls.empty:
        st.info("Aucune donnee disponible pour calculer les analyses avancees.")
        return

    total_calls = float(filtered_calls["record_count"].sum())
    resolved = float(filtered_calls.loc[filtered_calls["statut"] == "Resolu", "record_count"].sum())
    resolution_rate = (resolved / total_calls * 100.0) if total_calls > 0 else 0.0
    alert_volume = float(
        filtered_calls.loc[filtered_calls["categorie"].astype(str).str.contains("alerte", case=False, na=False), "record_count"].sum()
    )
    alert_rate = (alert_volume / total_calls * 100.0) if total_calls > 0 else 0.0

    if "Performance de resolution" in options:
        c1, c2, c3 = st.columns(3)
        c1.metric("Taux de resolution", f"{resolution_rate:.1f}%")
        c2.metric("Volume alerte (appels)", format_int(alert_volume))
        c3.metric("Part alerte", f"{alert_rate:.1f}%")

    if "Comparaison periode precedente" in options:
        prev_total = float(previous_calls["record_count"].sum()) if not previous_calls.empty else 0.0
        delta_pct = ((total_calls - prev_total) / prev_total * 100.0) if prev_total > 0 else np.nan
        window_days = (date_end.normalize() - date_start.normalize()).days + 1
        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Periode active", f"{window_days} jours")
        col_b.metric("Appels (periode active)", format_int(total_calls))
        if np.isnan(delta_pct):
            col_c.metric("Variation vs periode precedente", "N/A")
        else:
            col_c.metric("Variation vs periode precedente", f"{delta_pct:+.1f}%")

    if "Qualite des donnees (completude)" in options:
        comp_table, overall = compute_completeness_table(filtered_calls)
        comp_plot = comp_table.copy()
        comp_plot["Niveau"] = comp_plot["Completude (%)"].map(completeness_level)
        c1, c2 = st.columns([1.1, 2.2], gap="small")
        with c1:
            st.metric("Completude globale", f"{overall:.1f}%")
            weak = comp_table[comp_table["Completude (%)"] < 90.0]
            st.metric("Champs < 90%", int(len(weak)))
        with c2:
            fig_comp = px.bar(
                comp_plot.sort_values("Completude (%)", ascending=True),
                x="Completude (%)",
                y="Champ",
                orientation="h",
                range_x=[0, 100],
                color="Niveau",
                category_orders={"Niveau": ["< 15%", "15 - 49%", "50 - 90%", "> 90%"]},
                color_discrete_map={
                    "< 15%": "#dc2626",
                    "15 - 49%": "#f97316",
                    "50 - 90%": "#facc15",
                    "> 90%": "#16a34a",
                },
                hover_data={"Feuilles avec vides": True, "Niveau": False},
            )
            add_bar_value_labels(fig_comp, orientation="h", is_percent=True, expand_axis=False)
            fig_comp.update_layout(
                title="Completude des champs",
                height=280,
                margin=dict(l=0, r=0, t=40, b=0),
                legend_title_text="Seuil",
            )
            st.plotly_chart(fig_comp, use_container_width=True, key="analysis_completude_chart")
        feedback = build_missing_feedback_table(filtered_calls)
        if not feedback.empty:
            st.caption("Feedback qualite par fichier/feuille")
            st.dataframe(feedback, use_container_width=True, height=230, hide_index=True)

    if "Profil des appels" in options:
        row = st.columns(2)
        with row[0]:
            top_incidents = (
                filtered_calls.groupby("incident", as_index=False)["record_count"]
                .sum()
                .sort_values("record_count", ascending=False)
                .head(8)
            )
            fig_inc = px.bar(top_incidents, x="incident", y="record_count", labels={"record_count": "Record Count", "incident": "Incident"})
            add_bar_value_labels(fig_inc, orientation="v")
            fig_inc.update_layout(title="Top incidents", height=280, margin=dict(l=0, r=0, t=40, b=0))
            st.plotly_chart(fig_inc, use_container_width=True, key="analysis_top_incidents_chart")
        with row[1]:
            filtered_calls = filtered_calls.copy()
            filtered_calls["heure"] = pd.to_datetime(filtered_calls["date"], errors="coerce").dt.hour
            by_hour = filtered_calls.groupby("heure", as_index=False)["record_count"].sum().sort_values("heure")
            fig_hour = px.line(by_hour, x="heure", y="record_count", markers=True, labels={"heure": "Heure", "record_count": "Record Count"})
            add_line_value_labels(fig_hour)
            fig_hour.update_layout(title="Distribution horaire des appels", height=280, margin=dict(l=0, r=0, t=40, b=0))
            st.plotly_chart(fig_hour, use_container_width=True, key="analysis_hour_chart")


def render_general_page(filtered: pd.DataFrame, selected_provinces: list[str]) -> None:
    """Page infos generales avec bascule dynamique province -> territoire."""
    if filtered.empty:
        st.warning("Aucune donnee d'appels pour ces filtres.")
        return

    kpis = compute_kpis(filtered)
    total_calls = kpis["total_calls"]
    provinces_count = kpis["provinces_count"]
    selected_single_province = resolve_single_selected_province(filtered, selected_provinces)
    by_province = (
        filtered.groupby("province", as_index=False)["record_count"]
        .sum()
        .sort_values("record_count", ascending=True)
    )

    left_col, right_col = st.columns([2.7, 1.8], gap="small")

    with left_col:
        st.subheader("Nature/type d'appel")
        details = (
            filtered.groupby(["province", "territoire", "details"], as_index=False)["record_count"]
            .sum()
            .sort_values(["province", "record_count"], ascending=[True, False])
            .rename(columns={"province": "Province", "territoire": "Territoire", "details": "Details de l'appel", "record_count": "Record Count"})
        )
        details["Record Count"] = pd.to_numeric(details["Record Count"], errors="coerce").fillna(0).round().astype(int)
        details_preview = details.head(400)
        try:
            st.dataframe(
                details_preview.style.background_gradient(subset=["Record Count"], cmap="Reds"),
                use_container_width=True,
                height=470,
            )
        except Exception:
            # Streamlit Cloud may not have matplotlib; fallback to a plain dataframe.
            st.dataframe(
                details_preview,
                use_container_width=True,
                height=470,
            )

    with right_col:
        c1, c2 = st.columns(2)
        c1.metric("Nbre des provinces", format_int(provinces_count))
        c2.metric("Cumul des appels", format_int(total_calls))

        if selected_single_province:
            st.subheader(f"Proportion d'appel par territoire ({selected_single_province})")
            by_scope = (
                filtered.groupby("territoire", as_index=False)["record_count"]
                .sum()
                .sort_values("record_count", ascending=True)
            )
            y_col = "territoire"
            y_label = "Territoire"
        else:
            st.subheader("Proportion d'appel par province")
            by_scope = by_province
            y_col = "province"
            y_label = "Province"

        fig_prov = px.bar(
            by_scope,
            x="record_count",
            y=y_col,
            orientation="h",
            labels={"record_count": "Record Count", y_col: y_label},
            color="record_count",
            color_continuous_scale="Blues",
        )
        add_bar_value_labels(fig_prov, orientation="h")
        fig_prov.update_layout(
            title=f"Proportion d'appel par {y_label.lower()}" + (f" ({selected_single_province})" if selected_single_province else ""),
            margin=dict(l=10, r=10, t=40, b=10),
            coloraxis_showscale=False,
            height=265,
        )
        st.plotly_chart(fig_prov, use_container_width=True)

    st.subheader("Geolocalisation des appels par province")
    map_rows = []
    for _, row in by_province.iterrows():
        province = row["province"]
        if province in PROVINCE_COORDS:
            lat, lon = PROVINCE_COORDS[province]
            map_rows.append({"province": province, "lat": lat, "lon": lon, "calls": row["record_count"]})

    if map_rows:
        map_df = pd.DataFrame(map_rows)
        fig_map = px.scatter_mapbox(
            map_df,
            lat="lat",
            lon="lon",
            size="calls",
            color="calls",
            hover_name="province",
            hover_data={"calls": True, "lat": False, "lon": False},
            color_continuous_scale="Tealgrn",
            size_max=48,
            zoom=4.55,
            center={"lat": -3.5, "lon": 23.6},
            mapbox_style="carto-darkmatter",
        )
        fig_map.update_traces(text=map_df["calls"].map(format_int), mode="markers+text", textposition="top center")
        fig_map.update_layout(
            title="Geolocalisation des appels par province",
            margin=dict(l=0, r=0, t=40, b=0),
            height=560,
            coloraxis_showscale=False,
            mapbox=dict(pitch=0, bearing=0),
        )
        st.plotly_chart(fig_map, use_container_width=True)
    else:
        st.info("Ajoutez des noms de province reconnus pour afficher la carte.")

    st.subheader("Evolution du nombre d'appels au fil du temps")
    if selected_single_province:
        trend_col = "territoire"
        trend_label = "Territoire"
        top_scopes = (
            filtered.groupby("territoire", as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=False)
            .head(6)["territoire"]
            .tolist()
        )
        trend_source = filtered[filtered["territoire"].isin(top_scopes)].copy()
    else:
        trend_col = "province"
        trend_label = "Province"
        top_scopes = (
            filtered.groupby("province", as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=False)
            .head(6)["province"]
            .tolist()
        )
        trend_source = filtered[filtered["province"].isin(top_scopes)].copy()

    trend = group_by_day(trend_source, value_col="record_count", category_col=trend_col)
    if trend.empty:
        st.info("Aucune donnee de tendance disponible.")
        return
    trend = trend.sort_values(["date", trend_col]).copy()
    fig_trend = px.line(
        trend,
        x="date",
        y="record_count",
        color=trend_col,
        labels={"record_count": "Record Count", "date": "Date", trend_col: trend_label},
    )
    add_line_end_labels(fig_trend, trend, trend_col, value_col="record_count")
    fig_trend.update_traces(mode="lines+markers", marker_size=5, line_width=2.2)
    fig_trend.update_layout(
        title=f"Evolution du nombre d'appels au fil du temps par {trend_label.lower()}" + (f" ({selected_single_province})" if selected_single_province else ""),
        legend_orientation="h",
        legend_y=1.06,
        margin=dict(l=0, r=0, t=48, b=0),
        height=360,
        hovermode="x unified",
        xaxis_title="Date",
        yaxis_title="Record Count",
    )
    st.plotly_chart(fig_trend, use_container_width=True)


def render_details_page(filtered: pd.DataFrame, selected_provinces: list[str]) -> None:
    """Page autres details avec visuels complementaires (sans dupliquer la page generale)."""
    if filtered.empty:
        st.warning("Aucune donnee d'appels pour ces filtres.")
        return

    selected_single_province = resolve_single_selected_province(filtered, selected_provinces)
    data = filtered.copy()
    data["record_count"] = pd.to_numeric(data["record_count"], errors="coerce").fillna(0)

    top_row = st.columns(3, gap="small")

    with top_row[0]:
        st.subheader("Repartition par genre")
        by_gender = data.groupby("genre", as_index=False)["record_count"].sum()
        fig_gender = px.pie(
            by_gender,
            names="genre",
            values="record_count",
            hole=0.6,
            color="genre",
            color_discrete_map={"Homme": THEME["blue"], "Femme": THEME["pink"], "ND": "#94a3b8"},
        )
        fig_gender.update_traces(textinfo="label+percent+value")
        fig_gender.update_layout(title="Repartition par genre", margin=dict(l=0, r=0, t=40, b=0), height=310)
        st.plotly_chart(fig_gender, use_container_width=True, key="details_gender_chart")

    with top_row[1]:
        st.subheader("Top incidents/pathologies")
        by_incident = (
            data.groupby("incident", as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=True)
            .tail(12)
        )
        fig_incident = px.bar(
            by_incident,
            x="record_count",
            y="incident",
            orientation="h",
            labels={"record_count": "Record Count", "incident": "Incident/Pathologie"},
            color="record_count",
            color_continuous_scale="Blues",
        )
        add_bar_value_labels(fig_incident, orientation="h")
        fig_incident.update_layout(
            title="Top incidents/pathologies",
            margin=dict(l=0, r=0, t=40, b=0),
            coloraxis_showscale=False,
            height=310,
        )
        st.plotly_chart(fig_incident, use_container_width=True, key="details_incident_chart")

    with top_row[2]:
        if selected_single_province:
            st.subheader(f"Statut de resolution par territoire ({selected_single_province})")
            scope_col = "territoire"
            scope_label = "Territoire"
        else:
            st.subheader("Statut de resolution par province")
            scope_col = "province"
            scope_label = "Province"

        data[scope_col] = data[scope_col].fillna("Inconnu").astype(str).replace("", "Inconnu")
        top_scopes = (
            data.groupby(scope_col, as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=False)
            .head(12)[scope_col]
            .tolist()
        )
        by_status = (
            data[data[scope_col].isin(top_scopes)]
            .groupby([scope_col, "statut"], as_index=False)["record_count"]
            .sum()
        )
        status_order = ["Resolu", "Non resolu"]
        # Force l'affichage des deux statuts meme si l'un est absent dans les donnees.
        full_idx = pd.MultiIndex.from_product([top_scopes, status_order], names=[scope_col, "statut"])
        by_status = (
            by_status.set_index([scope_col, "statut"])
            .reindex(full_idx, fill_value=0)
            .reset_index()
        )
        zone_order = (
            by_status.groupby(scope_col, as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=True)[scope_col]
            .tolist()
        )
        fig_status = px.bar(
            by_status,
            x="record_count",
            y=scope_col,
            orientation="h",
            color="statut",
            barmode="group",
            category_orders={scope_col: zone_order, "statut": status_order},
            labels={"record_count": "Record Count", scope_col: scope_label, "statut": "Statut"},
            color_discrete_map={"Resolu": "#16a34a", "Non resolu": "#e61f25"},
        )
        add_bar_value_labels(fig_status, orientation="h")
        fig_status.update_layout(
            title=f"Statut de resolution par {scope_label.lower()}" + (f" ({selected_single_province})" if selected_single_province else ""),
            margin=dict(l=0, r=0, t=45, b=0),
            height=340,
            legend_title_text="Statut",
            legend_orientation="h",
            legend_y=1.08,
        )
        st.plotly_chart(fig_status, use_container_width=True, key="details_status_chart")

    second_row = st.columns(2, gap="small")

    with second_row[0]:
        scope_col = "territoire" if selected_single_province else "province"
        scope_label = "Territoire" if selected_single_province else "Province"
        if selected_single_province:
            st.subheader(f"Matrice {scope_label.lower()} x categorie ({selected_single_province})")
        else:
            st.subheader("Matrice province x categorie d'appel")

        top_scopes = (
            data.groupby(scope_col, as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=False)
            .head(10)[scope_col]
            .tolist()
        )
        top_categories = (
            data.groupby("categorie", as_index=False)["record_count"]
            .sum()
            .sort_values("record_count", ascending=False)
            .head(6)["categorie"]
            .tolist()
        )

        matrix_source = data[data[scope_col].isin(top_scopes) & data["categorie"].isin(top_categories)].copy()
        matrix = (
            matrix_source.groupby([scope_col, "categorie"], as_index=False)["record_count"]
            .sum()
            .pivot(index=scope_col, columns="categorie", values="record_count")
            .fillna(0)
        )

        if matrix.empty:
            st.info("Donnees insuffisantes pour la matrice.")
        else:
            fig_matrix = px.imshow(
                matrix,
                text_auto=".0f",
                aspect="auto",
                color_continuous_scale="Tealgrn",
                labels={"x": "Categorie d'appel", "y": scope_label, "color": "Record Count"},
            )
            fig_matrix.update_layout(
                title=("Matrice territoire x categorie" if selected_single_province else "Matrice province x categorie"),
                margin=dict(l=0, r=0, t=40, b=0),
                coloraxis_showscale=True,
                height=370,
            )
            st.plotly_chart(fig_matrix, use_container_width=True, key="details_matrix_chart")

    with second_row[1]:
        st.subheader("Appels par jour de la semaine")
        weekly = data.copy()
        weekly["date"] = pd.to_datetime(weekly["date"], errors="coerce")
        weekly = weekly.dropna(subset=["date"])

        if weekly.empty:
            st.info("Donnees de date indisponibles pour l'analyse hebdomadaire.")
        else:
            day_map = {
                0: "Lundi",
                1: "Mardi",
                2: "Mercredi",
                3: "Jeudi",
                4: "Vendredi",
                5: "Samedi",
                6: "Dimanche",
            }
            weekly["jour_num"] = weekly["date"].dt.dayofweek
            weekly["jour"] = weekly["jour_num"].map(day_map)
            by_day = (
                weekly.groupby(["jour_num", "jour"], as_index=False)["record_count"]
                .sum()
                .sort_values("jour_num")
            )
            fig_week = px.bar(
                by_day,
                x="jour",
                y="record_count",
                labels={"jour": "Jour", "record_count": "Record Count"},
                color="record_count",
                color_continuous_scale="Blues",
            )
            add_bar_value_labels(fig_week, orientation="v")
            fig_week.update_layout(
                title="Appels par jour de la semaine",
                margin=dict(l=0, r=0, t=40, b=0),
                coloraxis_showscale=False,
                height=370,
            )
            st.plotly_chart(fig_week, use_container_width=True, key="details_weekday_chart")


def render_alerts_page(
    alerts_df: pd.DataFrame,
    selected_provinces: list[str],
    calls_df_for_alerts: pd.DataFrame | None = None,
) -> None:
    """Page alertes detaillees avec tableaux et graphiques reactifs province -> territoire."""
    if alerts_df.empty:
        st.warning("Aucune donnee d'alertes disponible.")
        return

    selected_single_province = None
    selected_clean = [canonical_province(p) for p in selected_provinces if str(p).strip()]
    if len(selected_clean) == 1:
        selected_single_province = selected_clean[0]

    alerts_scope = alerts_df.copy()
    alerts_scope["territoire_norm"] = alerts_scope["location"].map(canonical_territory_name)
    prov_from_location = alerts_scope["location"].map(canonical_province)
    prov_from_territory = alerts_scope["territoire_norm"].map(province_from_territory)
    alerts_scope["province_scope"] = np.where(
        prov_from_location.ne("Inconnu"),
        prov_from_location,
        prov_from_territory.fillna("Inconnu"),
    )
    if selected_single_province:
        alerts_scope["location_scope"] = alerts_scope["territoire_norm"]
    else:
        alerts_scope["location_scope"] = alerts_scope["location"]

    # Base appels pour les visuels "categories des appels" et le tableau detaille des alertes.
    calls_scope = calls_df_for_alerts.copy() if isinstance(calls_df_for_alerts, pd.DataFrame) else pd.DataFrame()
    if not calls_scope.empty:
        calls_scope["province"] = calls_scope["province"].map(canonical_province)
        calls_scope["territoire"] = calls_scope["territoire"].map(canonical_territory_name)
        calls_scope["categorie"] = calls_scope["categorie"].fillna("Non classe").astype(str)
        calls_scope["incident"] = calls_scope["incident"].fillna("Non precise").astype(str)
        calls_scope["details"] = calls_scope["details"].fillna("Sans detail").astype(str)
        calls_scope["statut"] = calls_scope["statut"].fillna("Non resolu").astype(str)
        calls_scope["record_count"] = pd.to_numeric(calls_scope["record_count"], errors="coerce").fillna(0)

    if selected_single_province and not calls_scope.empty:
        locations = sorted(calls_scope["territoire"].dropna().astype(str).unique().tolist())
    else:
        locations = sorted(alerts_scope["location_scope"].dropna().astype(str).unique().tolist())
    indicators = sorted(alerts_scope["indicator"].dropna().astype(str).unique().tolist())

    st.markdown("<div class='filter-title'>Filtres alertes</div>", unsafe_allow_html=True)
    col_f1, col_f2 = st.columns(2)
    loc_label = "Territoire" if selected_single_province else "Localite"
    selected_locations = col_f1.multiselect(loc_label, options=locations, default=locations)
    selected_indicators = col_f2.multiselect("Indicateur", options=indicators, default=indicators)

    filtered = alerts_scope.copy()
    if selected_locations:
        if selected_single_province and not calls_scope.empty:
            # Les localites selectionnees pilotent principalement la vue APPELS
            # (territoires). On conserve les alertes filtrees par province.
            pass
        elif selected_single_province:
            filtered = filtered[
                filtered["territoire_norm"].isin(selected_locations) | filtered["location_scope"].isin(selected_locations)
            ]
        else:
            filtered = filtered[filtered["location_scope"].isin(selected_locations)]
    if selected_indicators:
        filtered = filtered[filtered["indicator"].isin(selected_indicators)]

    calls_scope_filtered = calls_scope.copy()
    if selected_single_province and not calls_scope_filtered.empty and selected_locations:
        calls_scope_filtered = calls_scope_filtered[calls_scope_filtered["territoire"].isin(selected_locations)].copy()

    if filtered.empty:
        st.info("Aucune alerte pour les filtres choisis.")
        return

    total_alerts = filtered["value"].sum()
    locations_count = filtered["location_scope"].nunique()
    indicators_count = filtered["indicator"].nunique()
    locations_metric_label = "Nb territoires" if selected_single_province else "Nb localites"
    if selected_single_province and not calls_scope_filtered.empty:
        metric_source = calls_scope_filtered[calls_scope_filtered["categorie"].str.contains("alerte", case=False, na=False)].copy()
        if metric_source.empty:
            metric_source = calls_scope_filtered.copy()
        total_alerts = metric_source["record_count"].sum()
        locations_count = metric_source["territoire"].nunique()

    c1, c2, c3 = st.columns(3)
    c1.metric("Total alertes", format_int(total_alerts))
    c2.metric(locations_metric_label, format_int(locations_count))
    c3.metric("Nb indicateurs", format_int(indicators_count))

    row = st.columns(2)
    with row[0]:
        if selected_single_province:
            st.subheader(f"Alertes par territoire ({selected_single_province})")
        else:
            st.subheader("Alertes par localite")

        # Si une province est selectionnee, on privilegie les alertes issues des APPELS
        # (categorie contenant "alerte") pour obtenir une vue par territoire coherente.
        by_location = pd.DataFrame()
        if selected_single_province and not calls_scope_filtered.empty:
            calls_alert_for_location = calls_scope_filtered[
                calls_scope_filtered["categorie"].str.contains("alerte", case=False, na=False)
            ].copy()
            if not calls_alert_for_location.empty:
                by_location = (
                    calls_alert_for_location.groupby("territoire", as_index=False)["record_count"]
                    .sum()
                    .rename(columns={"territoire": "location_scope", "record_count": "value"})
                    .sort_values("value", ascending=True)
                    .tail(20)
                )
        if by_location.empty:
            by_location = (
                filtered.groupby("location_scope", as_index=False)["value"]
                .sum()
                .sort_values("value", ascending=True)
                .tail(20)
            )

        fig_loc = px.bar(
            by_location,
            x="value",
            y="location_scope",
            orientation="h",
            labels={"value": "Record Count", "location_scope": loc_label},
            color="value",
            color_continuous_scale="Blues",
        )
        add_bar_value_labels(fig_loc, orientation="h")
        fig_loc.update_layout(
            title=f"Alertes par {loc_label.lower()}" + (f" ({selected_single_province})" if selected_single_province else ""),
            margin=dict(l=0, r=0, t=40, b=0),
            coloraxis_showscale=False,
            height=350,
        )
        st.plotly_chart(fig_loc, use_container_width=True, key="alerts_location_chart")

    with row[1]:
        st.subheader("Alertes par indicateur")
        by_indicator = (
            filtered.groupby("indicator", as_index=False)["value"]
            .sum()
            .sort_values("value", ascending=True)
            .tail(20)
        )
        if by_indicator.empty:
            st.info("Aucun indicateur disponible pour les filtres selectionnes.")
        else:
            fig_indicator = px.bar(
                by_indicator,
                x="value",
                y="indicator",
                orientation="h",
                labels={"value": "Record Count", "indicator": "Indicateur"},
                color="value",
                color_continuous_scale="Tealgrn",
            )
            add_bar_value_labels(fig_indicator, orientation="h")
            fig_indicator.update_layout(
                title="Alertes par indicateur",
                margin=dict(l=0, r=0, t=40, b=0),
                coloraxis_showscale=False,
                height=350,
            )
            st.plotly_chart(fig_indicator, use_container_width=True, key="alerts_indicator_chart")

    st.subheader("Evolution des alertes")
    if selected_single_province:
        if not calls_scope_filtered.empty:
            trend_calls = calls_scope_filtered[calls_scope_filtered["categorie"].str.contains("alerte", case=False, na=False)].copy()
            if trend_calls.empty:
                trend_calls = calls_scope_filtered.copy()
            top_territories = (
                trend_calls.groupby("territoire", as_index=False)["record_count"]
                .sum()
                .sort_values("record_count", ascending=False)
                .head(6)["territoire"]
                .tolist()
            )
            trend_source = trend_calls[trend_calls["territoire"].isin(top_territories)].copy()
            trend = group_by_day(trend_source, value_col="record_count", category_col="territoire")
            trend = trend.rename(columns={"record_count": "value", "territoire": "location_scope"})
        else:
            top_territories = (
                filtered.groupby("location_scope", as_index=False)["value"]
                .sum()
                .sort_values("value", ascending=False)
                .head(6)["location_scope"]
                .tolist()
            )
            trend_source = filtered[filtered["location_scope"].isin(top_territories)].copy()
            trend = group_by_day(trend_source, value_col="value", category_col="location_scope")
        trend_color_col = "location_scope"
        trend_labels = {"date": "Date", "value": "Record Count", "location_scope": "Territoire"}
    else:
        top_provinces = (
            filtered.groupby("province_scope", as_index=False)["value"]
            .sum()
            .sort_values("value", ascending=False)
            .head(6)["province_scope"]
            .tolist()
        )
        trend_source = filtered[filtered["province_scope"].isin(top_provinces)].copy()
        trend = group_by_day(trend_source, value_col="value", category_col="province_scope")
        trend_color_col = "province_scope"
        trend_labels = {"date": "Date", "value": "Record Count", "province_scope": "Province"}

    if trend.empty:
        st.info("Aucune donnee de tendance alertes disponible.")
        return
    trend = trend.sort_values(["date", trend_color_col]).copy()
    fig_alert_trend = px.line(
        trend,
        x="date",
        y="value",
        color=trend_color_col,
        labels=trend_labels,
    )
    add_line_end_labels(fig_alert_trend, trend, trend_color_col, value_col="value")
    fig_alert_trend.update_traces(mode="lines+markers", marker_size=6, line_width=2.5)
    fig_alert_trend.update_layout(
        title="Evolution des alertes",
        legend_orientation="h",
        legend_y=1.06,
        margin=dict(l=0, r=0, t=48, b=0),
        height=380,
        hovermode="x unified",
        xaxis_title="Date",
        yaxis_title="Record Count",
    )
    st.plotly_chart(fig_alert_trend, use_container_width=True, key="alerts_trend_chart")

    st.subheader("Tableau detaille des alertes")
    calls_alert_scope = calls_scope_filtered.copy()
    fallback_alert_scope = pd.DataFrame()

    # Dictionnaires de details textuels issus de la base ALERTES (pour remplacer
    # les colonnes "details" numeriques des APPELS quand c'est possible).
    alert_details_lookup = filtered.copy()
    alert_details_lookup["details_text"] = (
        alert_details_lookup["details"].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
    )
    text_mask = alert_details_lookup["details_text"].ne("") & ~alert_details_lookup["details_text"].str.fullmatch(
        r"[0-9\+\-\s\(\)\.]{6,}", na=False
    )
    alert_details_lookup = alert_details_lookup.loc[text_mask].copy()
    alert_details_lookup["province_key"] = alert_details_lookup["province_scope"].fillna("Inconnu").astype(str)
    alert_details_lookup["territoire_key"] = alert_details_lookup["territoire_norm"].fillna("Inconnu").astype(str)
    alert_details_lookup["pathologie_key"] = alert_details_lookup["indicator"].map(canonical_pathology_name).fillna("Alerte")

    if not alert_details_lookup.empty:
        lookup_province = (
            alert_details_lookup.groupby(["province_key", "pathologie_key"], as_index=False)["details_text"]
            .agg(lambda s: s.value_counts(dropna=True).index[0] if not s.empty else "")
            .rename(columns={"details_text": "details_lookup"})
        )
        lookup_territoire = (
            alert_details_lookup.groupby(["territoire_key", "pathologie_key"], as_index=False)["details_text"]
            .agg(lambda s: s.value_counts(dropna=True).index[0] if not s.empty else "")
            .rename(columns={"details_text": "details_lookup"})
        )
    else:
        lookup_province = pd.DataFrame(columns=["province_key", "pathologie_key", "details_lookup"])
        lookup_territoire = pd.DataFrame(columns=["territoire_key", "pathologie_key", "details_lookup"])

    if not calls_alert_scope.empty:
        alert_mask = calls_alert_scope["categorie"].str.contains("alerte", case=False, na=False)
        if alert_mask.any():
            calls_alert_scope = calls_alert_scope.loc[alert_mask].copy()

        calls_alert_scope["details"] = (
            calls_alert_scope["details"].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
        )
        # Si la colonne details contient surtout des numeros, on remplace en priorite
        # par un texte issu d'ALERTES (meme zone + pathologie), sinon libelle explicatif.
        if "details" in calls_alert_scope.columns and is_mostly_numeric_series(calls_alert_scope["details"]):
            calls_alert_scope["pathologie_key"] = calls_alert_scope["incident"].map(canonical_pathology_name).fillna("Non precise")
            raw_details = calls_alert_scope["details"].fillna("").astype(str).str.strip()
            numeric_mask = raw_details.eq("") | raw_details.str.fullmatch(r"[0-9\+\-\s\(\)\.]{6,}", na=False)

            if selected_single_province and not lookup_territoire.empty:
                calls_alert_scope["territoire_key"] = calls_alert_scope["territoire"].fillna("Inconnu").astype(str)
                calls_alert_scope = calls_alert_scope.merge(
                    lookup_territoire,
                    on=["territoire_key", "pathologie_key"],
                    how="left",
                )
            elif not selected_single_province and not lookup_province.empty:
                calls_alert_scope["province_key"] = calls_alert_scope["province"].fillna("Inconnu").astype(str)
                calls_alert_scope = calls_alert_scope.merge(
                    lookup_province,
                    on=["province_key", "pathologie_key"],
                    how="left",
                )
            else:
                calls_alert_scope["details_lookup"] = np.nan

            calls_alert_scope["details"] = np.where(
                ~numeric_mask,
                raw_details,
                calls_alert_scope["details_lookup"].fillna(
                    "L'appelant signale: " + calls_alert_scope["incident"].fillna("pathologie non precise").astype(str)
                ),
            )
            calls_alert_scope.drop(
                columns=["details_lookup", "pathologie_key", "territoire_key", "province_key"],
                errors="ignore",
                inplace=True,
            )
        calls_alert_scope.loc[calls_alert_scope["details"].eq(""), "details"] = "Sans detail"

    if calls_alert_scope.empty:
        # Fallback si les appels ne sont pas disponibles: on reconstruit le tableau depuis ALERTES.
        fallback = filtered.sort_values(["value", "date"], ascending=[False, False]).copy()
        fallback_alert_scope = fallback.copy()
        fallback["Province"] = fallback["province_scope"].fillna("Inconnu")
        fallback["Territoire"] = fallback["territoire_norm"].fillna("Inconnu")
        fallback["Pathologie"] = fallback["indicator"].map(canonical_pathology_name).fillna("Alerte")

        fallback["pathologie_key"] = fallback["Pathologie"].astype(str)
        base_fallback_details = fallback["details"].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
        fallback_numeric_mask = base_fallback_details.eq("") | base_fallback_details.str.fullmatch(
            r"[0-9\+\-\s\(\)\.]{6,}", na=False
        )

        if selected_single_province and not lookup_territoire.empty:
            fallback["territoire_key"] = fallback["Territoire"].fillna("Inconnu").astype(str)
            fallback = fallback.merge(
                lookup_territoire,
                on=["territoire_key", "pathologie_key"],
                how="left",
            )
        elif not selected_single_province and not lookup_province.empty:
            fallback["province_key"] = fallback["Province"].fillna("Inconnu").astype(str)
            fallback = fallback.merge(
                lookup_province,
                on=["province_key", "pathologie_key"],
                how="left",
            )
        else:
            fallback["details_lookup"] = np.nan

        fallback["Details de l'appel"] = np.where(
            ~fallback_numeric_mask,
            base_fallback_details,
            fallback["details_lookup"].fillna("L'appelant signale: " + fallback["Pathologie"].astype(str)),
        )
        fallback.loc[fallback["Details de l'appel"].eq(""), "Details de l'appel"] = "Sans detail"
        fallback.drop(
            columns=["details_lookup", "pathologie_key", "territoire_key", "province_key"],
            errors="ignore",
            inplace=True,
        )
        fallback["Resolution"] = "Non renseigne"
        if selected_single_province:
            table = fallback[["Territoire", "Pathologie", "Details de l'appel", "Resolution"]].sort_values(
                ["Territoire", "Pathologie", "Details de l'appel"]
            )
        else:
            table = fallback[["Province", "Pathologie", "Details de l'appel", "Resolution"]].sort_values(
                ["Province", "Pathologie", "Details de l'appel"]
            )
    else:
        calls_alert_scope = calls_alert_scope.sort_values(["province", "territoire", "incident", "date"]).copy()
        if selected_single_province:
            table = calls_alert_scope.rename(
                columns={
                    "territoire": "Territoire",
                    "incident": "Pathologie",
                    "details": "Details de l'appel",
                    "statut": "Resolution",
                }
            )[["Territoire", "Pathologie", "Details de l'appel", "Resolution"]]
        else:
            table = calls_alert_scope.rename(
                columns={
                    "province": "Province",
                    "incident": "Pathologie",
                    "details": "Details de l'appel",
                    "statut": "Resolution",
                }
            )[["Province", "Pathologie", "Details de l'appel", "Resolution"]]

    # Garantie finale: eviter que "Pathologie" et "Details de l'appel"
    # affichent exactement la meme valeur.
    table = normalize_call_details_text(table, path_col="Pathologie", details_col="Details de l'appel")
    table = table.fillna("Non renseigne").reset_index(drop=True)
    st.dataframe(table.head(800), use_container_width=True, height=360, hide_index=True)

    # Synthese graphique du tableau detaille: zones (province/territoire) par pathologie.
    if calls_alert_scope.empty:
        detail_source = fallback_alert_scope.copy()
        if detail_source.empty:
            st.info("Aucune donnee disponible pour la synthese.")
            return
        if selected_single_province:
            detail_source["zone"] = detail_source["territoire_norm"].fillna("Inconnu")
            zone_label = "Territoire"
        else:
            detail_source["zone"] = detail_source["province_scope"].fillna("Inconnu")
            zone_label = "Province"
        detail_source["pathologie"] = detail_source["indicator"].map(canonical_pathology_name).fillna("Alerte")
        detail_source["metric"] = pd.to_numeric(detail_source["value"], errors="coerce").fillna(0)
    else:
        if selected_single_province:
            detail_source = calls_alert_scope.rename(columns={"territoire": "zone", "incident": "pathologie", "record_count": "metric"})[
                ["zone", "pathologie", "metric"]
            ]
            zone_label = "Territoire"
        else:
            detail_source = calls_alert_scope.rename(columns={"province": "zone", "incident": "pathologie", "record_count": "metric"})[
                ["zone", "pathologie", "metric"]
            ]
            zone_label = "Province"
        detail_source["pathologie"] = detail_source["pathologie"].map(canonical_pathology_name).fillna("Non precise")

    detail_source["metric"] = pd.to_numeric(detail_source["metric"], errors="coerce").fillna(0)
    detail_source["zone"] = detail_source["zone"].fillna("Inconnu").astype(str)
    detail_source["pathologie"] = detail_source["pathologie"].fillna("Non precise").astype(str)
    top_zones = (
        detail_source.groupby("zone", as_index=False)["metric"]
        .sum()
        .sort_values("metric", ascending=False)
        .head(6)["zone"]
        .tolist()
    )
    top_pathologies = (
        detail_source.groupby("pathologie", as_index=False)["metric"]
        .sum()
        .sort_values("metric", ascending=False)
        .head(8)["pathologie"]
        .tolist()
    )
    detail_source = detail_source[
        detail_source["zone"].isin(top_zones) & detail_source["pathologie"].isin(top_pathologies)
    ].copy()
    detail_plot = (
        detail_source.groupby(["pathologie", "zone"], as_index=False)["metric"]
        .sum()
    )
    path_order = (
        detail_plot.groupby("pathologie", as_index=False)["metric"]
        .sum()
        .sort_values("metric", ascending=False)["pathologie"]
        .tolist()
    )

    if not detail_plot.empty:
        st.subheader("Synthese graphique du tableau detaille des alertes")
        fig_detail = px.bar(
            detail_plot,
            x="pathologie",
            y="metric",
            color="zone",
            barmode="group",
            category_orders={"pathologie": path_order},
            labels={"pathologie": "Pathologie", "metric": "Record Count", "zone": zone_label},
        )
        add_bar_value_labels(fig_detail, orientation="v")
        fig_detail.update_layout(
            title=f"Synthese des alertes par {zone_label.lower()} et pathologie"
            + (f" ({selected_single_province})" if selected_single_province else ""),
            height=430,
            margin=dict(l=0, r=0, t=42, b=0),
            legend_orientation="h",
            legend_y=1.08,
            xaxis_title="Pathologie",
            yaxis_title="Record Count",
        )
        fig_detail.update_xaxes(tickangle=-18)
        st.plotly_chart(fig_detail, use_container_width=True, key="alerts_details_synthesis_chart")


def default_postgres_config() -> PostgresConfig:
    # Priorite: champs auth saisis en sidebar (si presents), sinon secrets/env.
    auth_host = str(st.session_state.get("auth_pg_host", "")).strip()
    auth_port = str(st.session_state.get("auth_pg_port", "")).strip()
    auth_database = str(st.session_state.get("auth_pg_database", "")).strip()
    auth_user = str(st.session_state.get("auth_pg_user", "")).strip()
    auth_password = str(st.session_state.get("auth_pg_password", ""))
    auth_schema = str(st.session_state.get("auth_pg_schema", "")).strip()
    auth_sslmode = str(st.session_state.get("auth_pg_sslmode", "")).strip()

    return PostgresConfig(
        host=auth_host or get_secret_or_env("PGHOST", DEFAULT_DB_HOST),
        port=auth_port or get_secret_or_env("PGPORT", DEFAULT_DB_PORT),
        database=auth_database or get_secret_or_env("PGDATABASE", DEFAULT_DB_NAME),
        user=auth_user or get_secret_or_env("PGUSER", DEFAULT_DB_USER),
        password=auth_password or get_secret_or_env("PGPASSWORD", ""),
        schema=auth_schema or get_secret_or_env("PGSCHEMA", DEFAULT_DB_SCHEMA),
        sslmode=auth_sslmode or get_secret_or_env("PGSSLMODE", "prefer"),
    )


def render_postgres_sidebar(
    is_admin: bool = True,
    current_user: str = "",
) -> tuple[PostgresConfig, list[object] | None, str]:
    """Barre laterale PostgreSQL avec droits selon le role."""
    cfg = default_postgres_config()
    upload_files: list[object] | None = None
    write_mode = "append"

    with st.sidebar.expander("Connexion PostgreSQL", expanded=is_admin):
        if is_admin:
            st.caption("Renseignez la base PostgreSQL utilisee comme source de donnees du dashboard.")
            host = st.text_input("Host", value=cfg.host, key="pg_host")
            port = st.text_input("Port", value=cfg.port, key="pg_port")
            database = st.text_input("Base", value=cfg.database, key="pg_database")
            user = st.text_input("Utilisateur", value=cfg.user, key="pg_user")
            password = st.text_input("Mot de passe", value=cfg.password, type="password", key="pg_password")
            schema = st.text_input("Schema", value=cfg.schema, key="pg_schema")
            sslmode = st.selectbox(
                "SSL mode",
                options=["prefer", "disable", "require", "verify-ca", "verify-full"],
                index=["prefer", "disable", "require", "verify-ca", "verify-full"].index(cfg.sslmode)
                if cfg.sslmode in {"prefer", "disable", "require", "verify-ca", "verify-full"}
                else 0,
                key="pg_sslmode",
            )

            config = PostgresConfig(
                host=host.strip() or DEFAULT_DB_HOST,
                port=port.strip() or DEFAULT_DB_PORT,
                database=database.strip() or DEFAULT_DB_NAME,
                user=user.strip() or DEFAULT_DB_USER,
                password=password,
                schema=sanitize_identifier(schema.strip() or DEFAULT_DB_SCHEMA, DEFAULT_DB_SCHEMA),
                sslmode=sslmode,
            )

            if st.button("Tester la connexion", key="pg_test_connection"):
                try:
                    ok, db_message = ensure_postgres_database(config)
                    if not ok:
                        raise RuntimeError(db_message)
                    conn_url = build_pg_url(config)
                    schema_name = ensure_postgres_tables(conn_url, config.schema)
                    if db_message:
                        st.info(db_message)
                    st.success(f"Connexion OK. Schema actif: {schema_name}")
                except Exception as exc:
                    st.error(f"Echec connexion PostgreSQL: {exc}")
        else:
            config = cfg
            st.caption("Profil utilisateur: connexion en lecture seule (configuration reservee a l'administrateur).")
            st.text(f"Host: {cfg.host}")
            st.text(f"Port: {cfg.port}")
            st.text(f"Base: {cfg.database}")
            st.text(f"Schema: {cfg.schema}")

    if is_admin:
        with st.sidebar.expander("Importer Excel vers PostgreSQL", expanded=False):
            st.caption(
                "Import multiple autorise. Un seul uploader accepte tous les fichiers Excel, "
                "puis normalise et charge vers la table unique."
            )
            allow_import = render_admin_import_guard()
            write_mode = st.radio(
                "Mode d'ecriture",
                options=["append", "replace"],
                format_func=lambda x: "Ajouter (append)" if x == "append" else "Remplacer (replace)",
                horizontal=True,
                key="pg_write_mode",
                disabled=not allow_import,
            )
            upload_files = st.file_uploader(
                "Fichiers Excel a importer (APPELS/ALERTES melanges)",
                type=["xls", "xlsx"],
                key="pg_import_files",
                accept_multiple_files=True,
                disabled=not allow_import,
            )
            if st.button("Importer vers PostgreSQL", key="pg_import_button", disabled=not allow_import):
                try:
                    ok, db_message = ensure_postgres_database(config)
                    if not ok:
                        raise RuntimeError(db_message)
                    summary, report_df = import_uploaded_excels_to_postgres(
                        conn_url=build_pg_url(config),
                        schema=config.schema,
                        excel_files=upload_files,
                        write_mode=write_mode,
                    )
                    if db_message:
                        st.info(db_message)
                    st.session_state["pg_last_import_summary"] = summary
                    st.session_state["pg_last_import_report"] = report_df
                    st.success(summary)
                    st.rerun()
                except Exception as exc:
                    st.error(f"Import impossible: {exc}")

            last_summary = st.session_state.get("pg_last_import_summary", "")
            if last_summary:
                st.caption(last_summary)
            last_report = st.session_state.get("pg_last_import_report")
            if isinstance(last_report, pd.DataFrame) and not last_report.empty:
                report_view = last_report.copy()
                for col in ["total_rows", "rows_inserted", "duplicate_rows", "missing_rows"]:
                    if col in report_view.columns:
                        report_view[col] = pd.to_numeric(report_view[col], errors="coerce").fillna(0).astype(int)
                report_view = report_view.rename(
                    columns={
                        "dataset_type": "Jeu",
                        "file_name": "Fichier",
                        "sheet_name": "Feuille",
                        "date_min": "Date min",
                        "date_max": "Date max",
                        "total_rows": "Lignes lues",
                        "rows_inserted": "Lignes inserees",
                        "duplicate_rows": "Doublons",
                        "missing_columns": "Colonnes manquantes",
                        "missing_rows": "Lignes incompletes",
                        "status": "Statut",
                        "message": "Message",
                    }
                )
                st.dataframe(
                    report_view[
                        [
                            "Jeu",
                            "Fichier",
                            "Feuille",
                            "Date min",
                            "Date max",
                            "Lignes lues",
                            "Lignes inserees",
                            "Doublons",
                            "Colonnes manquantes",
                            "Lignes incompletes",
                            "Statut",
                            "Message",
                        ]
                    ],
                    use_container_width=True,
                    height=280,
                )
    else:
        with st.sidebar.expander("Importer Excel vers PostgreSQL", expanded=False):
            st.info("Fonction reservee a l'administrateur.")

    if is_admin:
        render_dashboard_users_admin(cfg=config, current_user=current_user)

    return config, upload_files, write_mode


def build_dashboard_report_excel(
    page: str,
    start_ts: pd.Timestamp,
    end_ts: pd.Timestamp,
    selected_provinces: list[str],
    selected_genres: list[str],
    selected_incidents: list[str],
    selected_categories: list[str],
    selected_years: list[int],
    selected_months: list[int],
    selected_weeks: list[int],
    selected_days: list[object],
    filtered_calls: pd.DataFrame,
    filtered_alerts: pd.DataFrame,
) -> bytes:
    """Genere un rapport Excel reatif aux filtres actifs et aux donnees affichees."""

    def fmt_values(values: list[object], formatter=None) -> str:
        if not values:
            return "Tous"
        if formatter is None:
            return ", ".join([str(v) for v in values])
        return ", ".join([formatter(v) for v in values])

    def safe_sheet_name(name: str, used: set[str]) -> str:
        cleaned = re.sub(r"[\[\]\*:/\\\?]", "_", str(name)).strip() or "Feuille"
        base = cleaned[:31]
        candidate = base
        i = 1
        while candidate in used:
            suffix = f"_{i}"
            candidate = (base[: max(1, 31 - len(suffix))] + suffix)[:31]
            i += 1
        used.add(candidate)
        return candidate

    selected_clean = [canonical_province(p) for p in selected_provinces if str(p).strip()]
    selected_single_province = selected_clean[0] if len(selected_clean) == 1 else None

    calls = filtered_calls.copy() if isinstance(filtered_calls, pd.DataFrame) else empty_calls_dataframe()
    alerts = filtered_alerts.copy() if isinstance(filtered_alerts, pd.DataFrame) else empty_alerts_dataframe()
    if not calls.empty:
        calls["record_count"] = pd.to_numeric(calls["record_count"], errors="coerce").fillna(0)
        calls["incident"] = calls["incident"].map(canonical_pathology_name)

    # Scope alertes harmonise (province -> territoire).
    alerts_scope = alerts.copy()
    if not alerts_scope.empty:
        alerts_scope["territoire_norm"] = alerts_scope["location"].map(canonical_territory_name)
        prov_from_location = alerts_scope["location"].map(canonical_province)
        prov_from_territory = alerts_scope["territoire_norm"].map(province_from_territory)
        alerts_scope["province_scope"] = np.where(
            prov_from_location.ne("Inconnu"),
            prov_from_location,
            prov_from_territory.fillna("Inconnu"),
        )
        alerts_scope["location_scope"] = alerts_scope["territoire_norm"] if selected_single_province else alerts_scope["location"]
        alerts_scope["indicator"] = alerts_scope["indicator"].map(canonical_pathology_name)
        alerts_scope["value"] = pd.to_numeric(alerts_scope["value"], errors="coerce").fillna(0)

    # KPI principaux.
    if not calls.empty:
        calls_kpi = compute_kpis(calls)
    else:
        calls_kpi = {
            "provinces_count": 0,
            "total_calls": 0,
            "resolved_calls": 0,
            "unresolved_calls": 0,
            "male_calls": 0,
            "female_calls": 0,
            "nd_calls": 0,
        }
    total_alerts = float(alerts_scope["value"].sum()) if not alerts_scope.empty else 0.0
    nb_localites = int(alerts_scope["location_scope"].nunique()) if not alerts_scope.empty else 0
    nb_indicateurs = int(alerts_scope["indicator"].nunique()) if not alerts_scope.empty else 0

    resume_df = pd.DataFrame(
        [
            {"Indicateur": "Section active", "Valeur": page},
            {"Indicateur": "Periode debut", "Valeur": start_ts.strftime("%Y-%m-%d")},
            {"Indicateur": "Periode fin", "Valeur": end_ts.strftime("%Y-%m-%d")},
            {"Indicateur": "Total appels", "Valeur": int(calls_kpi["total_calls"])},
            {"Indicateur": "Appels resolus", "Valeur": int(calls_kpi["resolved_calls"])},
            {"Indicateur": "Appels non resolus", "Valeur": int(calls_kpi["unresolved_calls"])},
            {"Indicateur": "Total alertes", "Valeur": int(round(total_alerts))},
            {"Indicateur": "Nb localites/territoires (alertes)", "Valeur": nb_localites},
            {"Indicateur": "Nb indicateurs (alertes)", "Valeur": nb_indicateurs},
        ]
    )

    filtres_df = pd.DataFrame(
        [
            {"Filtre": "Province", "Valeur": fmt_values(selected_provinces)},
            {"Filtre": "Genre", "Valeur": fmt_values(selected_genres)},
            {"Filtre": "Incident/Pathologie", "Valeur": fmt_values(selected_incidents)},
            {"Filtre": "Categorie d'appel", "Valeur": fmt_values(selected_categories)},
            {"Filtre": "Annee", "Valeur": fmt_values(selected_years)},
            {"Filtre": "Mois", "Valeur": fmt_values(selected_months, format_month_option)},
            {"Filtre": "Semaine (ISO)", "Valeur": fmt_values(selected_weeks)},
            {"Filtre": "Jour", "Valeur": fmt_values(selected_days, lambda d: pd.Timestamp(d).strftime("%Y-%m-%d"))},
            {"Filtre": "Periode active", "Valeur": f"{start_ts.strftime('%Y-%m-%d')} -> {end_ts.strftime('%Y-%m-%d')}"},
        ]
    )

    # Donnees source des graphiques appels.
    if calls.empty:
        proportion_calls_df = pd.DataFrame(columns=["Zone", "Record Count"])
        evol_calls_df = pd.DataFrame(columns=["Date", "Zone", "Record Count"])
        calls_alert_scope = pd.DataFrame(columns=calls.columns.tolist())
    else:
        zone_col = "territoire" if selected_single_province else "province"
        proportion_calls_df = (
            calls.groupby(zone_col, as_index=False)["record_count"]
            .sum()
            .rename(columns={zone_col: "Zone", "record_count": "Record Count"})
            .sort_values("Record Count", ascending=False)
        )
        top_zones = proportion_calls_df.head(6)["Zone"].tolist()
        evol_calls_df = group_by_day(
            calls[calls[zone_col].isin(top_zones)].copy(),
            value_col="record_count",
            category_col=zone_col,
        ).rename(columns={"date": "Date", zone_col: "Zone", "record_count": "Record Count"})

        calls_alert_scope = calls[calls["categorie"].astype(str).str.contains("alerte", case=False, na=False)].copy()
        if calls_alert_scope.empty:
            calls_alert_scope = calls.copy()

    # Donnees source des graphiques alertes.
    if alerts_scope.empty:
        alertes_localite_df = pd.DataFrame(columns=["Zone", "Record Count"])
        alertes_indicateur_df = pd.DataFrame(columns=["Indicateur", "Record Count"])
        evol_alertes_df = pd.DataFrame(columns=["Date", "Zone", "Record Count"])
    else:
        alertes_localite_df = (
            alerts_scope.groupby("location_scope", as_index=False)["value"]
            .sum()
            .rename(columns={"location_scope": "Zone", "value": "Record Count"})
            .sort_values("Record Count", ascending=False)
        )
        alertes_indicateur_df = (
            alerts_scope.groupby("indicator", as_index=False)["value"]
            .sum()
            .rename(columns={"indicator": "Indicateur", "value": "Record Count"})
            .sort_values("Record Count", ascending=False)
        )
        trend_col = "location_scope" if selected_single_province else "province_scope"
        top_alert_zones = (
            alerts_scope.groupby(trend_col, as_index=False)["value"]
            .sum()
            .sort_values("value", ascending=False)
            .head(6)[trend_col]
            .tolist()
        )
        evol_alertes_df = group_by_day(
            alerts_scope[alerts_scope[trend_col].isin(top_alert_zones)].copy(),
            value_col="value",
            category_col=trend_col,
        ).rename(columns={"date": "Date", trend_col: "Zone", "value": "Record Count"})

    # Tableau detaille des alertes (export).
    if not calls_alert_scope.empty:
        if selected_single_province:
            detail_alertes_df = calls_alert_scope.rename(
                columns={
                    "territoire": "Zone",
                    "incident": "Pathologie",
                    "details": "Details de l'appel",
                    "statut": "Resolution",
                    "date": "Date",
                    "record_count": "Record Count",
                }
            )[["Date", "Zone", "Pathologie", "Details de l'appel", "Resolution", "Record Count"]]
        else:
            detail_alertes_df = calls_alert_scope.rename(
                columns={
                    "province": "Zone",
                    "incident": "Pathologie",
                    "details": "Details de l'appel",
                    "statut": "Resolution",
                    "date": "Date",
                    "record_count": "Record Count",
                }
            )[["Date", "Zone", "Pathologie", "Details de l'appel", "Resolution", "Record Count"]]
        detail_alertes_df["Pathologie"] = detail_alertes_df["Pathologie"].map(canonical_pathology_name)
    else:
        if selected_single_province:
            detail_alertes_df = alerts_scope.rename(
                columns={
                    "date": "Date",
                    "territoire_norm": "Zone",
                    "indicator": "Pathologie",
                    "details": "Details de l'appel",
                    "value": "Record Count",
                }
            )[["Date", "Zone", "Pathologie", "Details de l'appel", "Record Count"]]
        else:
            detail_alertes_df = alerts_scope.rename(
                columns={
                    "date": "Date",
                    "province_scope": "Zone",
                    "indicator": "Pathologie",
                    "details": "Details de l'appel",
                    "value": "Record Count",
                }
            )[["Date", "Zone", "Pathologie", "Details de l'appel", "Record Count"]]
        detail_alertes_df["Resolution"] = "Non renseigne"
        detail_alertes_df = detail_alertes_df[["Date", "Zone", "Pathologie", "Details de l'appel", "Resolution", "Record Count"]]

    detail_alertes_df = normalize_call_details_text(detail_alertes_df, path_col="Pathologie", details_col="Details de l'appel")
    detail_alertes_df["Record Count"] = pd.to_numeric(detail_alertes_df["Record Count"], errors="coerce").fillna(0).round().astype(int)
    detail_alertes_df = detail_alertes_df.sort_values(["Zone", "Pathologie", "Date"], ascending=[True, True, False])

    # Synthese alerte par zone et pathologie.
    synth_alertes_df = (
        detail_alertes_df.groupby(["Zone", "Pathologie"], as_index=False)["Record Count"]
        .sum()
        .sort_values("Record Count", ascending=False)
    )

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Graph_Proportion_Appels", proportion_calls_df),
        ("Graph_Alertes_Localite", alertes_localite_df),
        ("Graph_Alertes_Indicateur", alertes_indicateur_df),
        ("Synthese_Alertes_Patho", synth_alertes_df),
        ("Appels_Filtres", calls),
        ("Alertes_Filtres", alerts_scope),
    ]

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        used_sheet_names: set[str] = set()
        for name, df in sheets:
            safe_name = safe_sheet_name(name, used_sheet_names)
            df_to_write = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
            if not df_to_write.empty:
                for col in df_to_write.columns:
                    if "date" in str(col).lower():
                        try:
                            df_to_write[col] = pd.to_datetime(df_to_write[col], errors="ignore")
                        except Exception:
                            pass
            df_to_write.to_excel(writer, sheet_name=safe_name, index=False)

        # Mise en forme "presentation" du classeur.
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill("solid", fgColor="0B5F7A")
            header_font = Font(color="FFFFFF", bold=True)

            for ws in writer.book.worksheets:
                ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    ws.auto_filter.ref = ws.dimensions
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                scan_max_row = min(ws.max_row, 500)
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for row in range(1, scan_max_row + 1):
                        val = ws.cell(row=row, column=col_idx).value
                        if val is None:
                            continue
                        max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)
        except Exception:
            # En cas d'absence de dependances de style, le fichier reste exportable.
            pass

    output.seek(0)
    return output.getvalue()


def main() -> None:
    """Point d'entree principal de l'application."""
    inject_styles()

    auth_ok, auth_user, auth_role = render_auth_sidebar()
    if not auth_ok:
        st.stop()

    is_admin = auth_role == "administrateur"
    st.sidebar.title("Dashboard call center 2025")
    st.sidebar.caption(f"Profil: {auth_user} ({auth_role})")
    page = st.sidebar.radio(
        "Section",
        ["Informations generales", "Autres details d'informations", "Details alertes"],
        index=0,
    )
    if is_admin:
        source_mode = st.sidebar.radio(
            "Source des donnees",
            ["PostgreSQL", "Upload Excel direct"],
            index=0,
            help="PostgreSQL: lit les tables SQL. Upload Excel direct: utilise uniquement les fichiers charges dans cette session.",
        )
    else:
        source_mode = "PostgreSQL"
        st.sidebar.info("Utilisateur: mode lecture seule (visualisation + export).")

    if is_admin and source_mode != "PostgreSQL":
        render_dashboard_users_admin(cfg=default_postgres_config(), current_user=auth_user)

    calls_df = empty_calls_dataframe()
    alerts_df = empty_alerts_dataframe()
    calls_info = DataSourceInfo("Aucune source", "-", "")
    alerts_info = DataSourceInfo("Aucune source", "-", "")

    if source_mode == "PostgreSQL":
        pg_config, _, _ = render_postgres_sidebar(is_admin=is_admin, current_user=auth_user)
        conn_url = build_pg_url(pg_config)
        db_label = f"PostgreSQL {pg_config.host}:{pg_config.port}/{pg_config.database}"
        try:
            ok, db_message = ensure_postgres_database(pg_config)
            if not ok:
                raise RuntimeError(db_message)
            if db_message:
                st.sidebar.success(db_message)
            calls_df, calls_info, alerts_df, alerts_info = load_postgres_data(
                conn_url=conn_url,
                schema=pg_config.schema,
                db_label=db_label,
            )
        except Exception as exc:
            render_header()
            st.error(f"Connexion PostgreSQL impossible: {exc}")
            st.stop()
    else:
        with st.sidebar.expander("Importer les fichiers Excel", expanded=True):
            st.caption("Chargez un ou plusieurs fichiers Excel. Le type est detecte automatiquement.")
            upload_files = st.file_uploader(
                "Fichiers Excel (.xls/.xlsx)",
                type=["xls", "xlsx"],
                key="upload_files",
                accept_multiple_files=True,
            )
        calls_df, calls_info, alerts_df, alerts_info = load_unified_data(upload_files)

    render_header()
    render_source_notes(calls_info, alerts_info)

    if source_mode == "Upload Excel direct":
        has_upload = bool(st.session_state.get("upload_files"))
        if not has_upload:
            st.info("Importez au moins un fichier Excel dans la barre laterale pour continuer.")
            st.stop()

    if calls_df.empty and alerts_df.empty:
        st.info("Aucune donnee disponible. Importez vos fichiers Excel vers PostgreSQL, puis rechargez la page.")
        st.stop()

    if calls_df.empty and page != "Details alertes":
        st.warning("La source APPELS est vide. Importez des appels pour afficher cette section.")
        st.stop()

    if calls_df.empty and page == "Details alertes":
        render_alerts_page(alerts_df, [], empty_calls_dataframe())
        st.stop()

    kpi_placeholder = st.empty()

    min_date = pd.to_datetime(calls_df["date"], errors="coerce").min()
    max_date = pd.to_datetime(calls_df["date"], errors="coerce").max()
    if pd.isna(min_date) or pd.isna(max_date):
        min_date = pd.Timestamp.today().normalize()
        max_date = min_date

    with st.container(border=True):
        st.markdown("<div class='filter-title'>Zone de filtres principaux</div>", unsafe_allow_html=True)
        # Bloc 1: filtres metier reactifs (province -> genre -> incident -> categorie).
        business_cols = st.columns(4, gap="small")
        province_options = sorted(calls_df["province"].dropna().astype(str).unique().tolist())
        selected_provinces = business_cols[0].multiselect("Province", province_options, default=province_options)

        genre_scope = calls_df.copy()
        if selected_provinces:
            genre_scope = genre_scope[genre_scope["province"].isin(selected_provinces)]
        genre_options = sorted(genre_scope["genre"].dropna().astype(str).unique().tolist())
        selected_genres = business_cols[1].multiselect("Genre", genre_options, default=genre_options)

        incident_scope = genre_scope.copy()
        if selected_genres:
            incident_scope = incident_scope[incident_scope["genre"].isin(selected_genres)]
        incident_options = sorted(incident_scope["incident"].dropna().astype(str).unique().tolist())
        selected_incidents = business_cols[2].multiselect("Incident/Pathologie", incident_options, default=incident_options)

        category_scope = incident_scope.copy()
        if selected_incidents:
            category_scope = category_scope[category_scope["incident"].isin(selected_incidents)]
        category_options = sorted(category_scope["categorie"].dropna().astype(str).unique().tolist())
        selected_categories = business_cols[3].multiselect("Categorie d'appel", category_options, default=category_options)

        # Bloc 2: granularite temporelle reactive.
        temporal_scope = category_scope.copy()
        if selected_categories:
            temporal_scope = temporal_scope[temporal_scope["categorie"].isin(selected_categories)]
        temporal_frame = build_temporal_filter_frame(temporal_scope)

        time_cols = st.columns(5, gap="small")
        year_options = sorted(temporal_frame["annee"].dropna().astype(int).unique().tolist())
        selected_years = time_cols[0].multiselect("Annee", options=year_options, default=year_options)

        month_frame = temporal_frame.copy()
        if selected_years:
            month_frame = month_frame[month_frame["annee"].isin(selected_years)]
        month_options = sorted(month_frame["mois"].dropna().astype(int).unique().tolist())
        selected_months = time_cols[1].multiselect(
            "Mois",
            options=month_options,
            default=month_options,
            format_func=format_month_option,
        )

        week_frame = month_frame.copy()
        if selected_months:
            week_frame = week_frame[week_frame["mois"].isin(selected_months)]
        week_options = sorted(week_frame["semaine"].dropna().astype(int).unique().tolist())
        selected_weeks = time_cols[2].multiselect("Semaine (ISO)", options=week_options, default=week_options)

        day_frame = week_frame.copy()
        if selected_weeks:
            day_frame = day_frame[day_frame["semaine"].isin(selected_weeks)]
        day_options = sorted(day_frame["jour"].dt.date.unique().tolist()) if not day_frame.empty else []
        selected_days = time_cols[3].multiselect(
            "Jour",
            options=day_options,
            default=day_options,
            format_func=lambda d: pd.Timestamp(d).strftime("%Y-%m-%d"),
        )

        # La periode depend explicitement de la selection annee/mois/semaine/jour.
        period_scope = day_frame.copy()
        if selected_days:
            period_scope = period_scope[period_scope["jour"].dt.date.isin(selected_days)]
        if period_scope.empty:
            period_scope = temporal_frame.copy()

        if period_scope.empty:
            period_min = min_date.date()
            period_max = max_date.date()
        else:
            period_min = pd.to_datetime(period_scope["jour"], errors="coerce").min().date()
            period_max = pd.to_datetime(period_scope["jour"], errors="coerce").max().date()

        date_selected = time_cols[4].date_input(
            "Selectionner la periode",
            value=(period_min, period_max),
            min_value=period_min,
            max_value=period_max,
        )

        if isinstance(date_selected, tuple) and len(date_selected) == 2:
            date_start, date_end = date_selected
        else:
            date_start = date_end = date_selected
        start_ts = pd.Timestamp(date_start)
        end_ts = pd.Timestamp(date_end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    filtered_calls = apply_calls_filters(
        calls_df,
        start_ts,
        end_ts,
        selected_provinces,
        selected_genres,
        selected_incidents,
        selected_categories,
        years=selected_years,
        months=selected_months,
        iso_weeks=selected_weeks,
        days=selected_days,
    )

    # KPI horizontal dynamique (mis a jour selon les filtres actifs)
    with kpi_placeholder.container():
        render_kpi_horizontal(compute_kpis(filtered_calls))

    window_days = (end_ts.normalize() - start_ts.normalize()).days + 1
    prev_end = start_ts - pd.Timedelta(seconds=1)
    prev_start = prev_end - pd.Timedelta(days=window_days) + pd.Timedelta(seconds=1)

    previous_calls = apply_calls_filters(
        calls_df,
        prev_start,
        prev_end,
        selected_provinces,
        selected_genres,
        selected_incidents,
        selected_categories,
    )
    filtered_alerts = apply_alerts_filters(
        alerts_df,
        start_ts,
        end_ts,
        selected_provinces,
        years=selected_years,
        months=selected_months,
        iso_weeks=selected_weeks,
        days=selected_days,
    )

    with st.sidebar.expander("Rapport final exportable (Excel)", expanded=False):
        st.caption(
            "Export reactif aux filtres actifs et aux manipulations du tableau de bord."
        )
        st.write(
            f"Periode: `{start_ts.strftime('%Y-%m-%d')}` -> `{end_ts.strftime('%Y-%m-%d')}`"
        )
        try:
            report_bytes = build_dashboard_report_excel(
                page=page,
                start_ts=start_ts,
                end_ts=end_ts,
                selected_provinces=selected_provinces,
                selected_genres=selected_genres,
                selected_incidents=selected_incidents,
                selected_categories=selected_categories,
                selected_years=selected_years,
                selected_months=selected_months,
                selected_weeks=selected_weeks,
                selected_days=selected_days,
                filtered_calls=filtered_calls,
                filtered_alerts=filtered_alerts,
            )
            st.download_button(
                "Telecharger le rapport Excel",
                data=report_bytes,
                file_name=f"rapport_call_center_{start_ts.strftime('%Y%m%d')}_{end_ts.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_dashboard_report_excel",
            )
        except Exception as exc:
            st.error(f"Generation du rapport impossible: {exc}")

    render_interactive_analytics(filtered_calls, previous_calls, start_ts, end_ts)

    if page == "Informations generales":
        render_general_page(filtered_calls, selected_provinces)
    elif page == "Autres details d'informations":
        render_details_page(filtered_calls, selected_provinces)
    else:
        render_alerts_page(filtered_alerts, selected_provinces, filtered_calls)


if __name__ == "__main__":
    main()
